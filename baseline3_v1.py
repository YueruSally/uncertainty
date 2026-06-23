#!/usr/bin/env python
# -*- coding: utf-8 -*-

import math
import random
import pathlib
import time
import argparse
from copy import deepcopy
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional
from pathlib import Path as FSPath
import json as _json

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401


# ════════════════════════════════════════════════════════
# Global settings
# ════════════════════════════════════════════════════════

TIME_BUCKET_H = 24.0

CHINA_REGIONS   = {"CN"}
EUROPE_REGIONS  = {"EE", "WE"}
TRANSIT_REGIONS = {"KZ", "KG", "UZ", "RU", "BY"}

CORRIDOR_ORDER: Dict[str, int] = {"CN": 0, "CA": 1, "RU": 2, "EE": 3, "WE": 4}

CHINA_BORDER_NODES: set = {"Erenhot", "Manzhouli", "Khorgos", "Lianyungang",
                            "Chongqing", "Yiwu"}
NODE_GROUP: Dict[str, str] = {}

# ════════════════════════════════════════════════════════
# [V1 WIRING] Congestion-node classification
# ────────────────────────────────────────────────────────
# NOTE: CHINA_BORDER_NODES above is used ONLY for routing
# monotonicity (china_border_monotone_ok) — do NOT repurpose it.
# The sets below control where border delay & capacity apply.
#
# Break-of-gauge / transit border nodes = the true 1435↔1520mm
# change points + shared chokepoint. ONLY these get fixed border
# delay AND a node capacity soft-constraint. Ordinary EU
# destinations (Berlin, Hamburg, Duisburg, Rotterdam, Prague…)
# are deliberately EXCLUDED even though the data generator wrongly
# tagged them IsBorderNode=1 with 40.6h delay + capacity 500.
BREAK_OF_GAUGE_NODES: set = {
    "Khorgos", "Alashankou", "Dostyk", "Altynkol",
    "Manzhouli", "Zabaykalsk", "Erenhot",
    "Brest", "Malaszewicze",
}
# Sea ports — candidate congestion nodes. Wired but only active
# when a positive BorderCapacity_TEUday is supplied for them
# (current data has none, so they stay inactive in v1).
PORT_CONGESTION_NODES: set = {
    "Shanghai", "Ningbo", "Tianjin", "Qingdao", "Dalian",
    "Rotterdam", "Hamburg", "Antwerp",
}
BOTTLENECK_NODES: set = set(BREAK_OF_GAUGE_NODES) | set(PORT_CONGESTION_NODES)

# Filled at load time from Node_Border (BorderCapacity / BackgroundFlow)
BORDER_CAPACITY: Dict[str, float] = {}
BACKGROUND_FLOW: Dict[str, float] = {}

# Soft-penalty weight for exceeding node/border daily capacity
PEN_BORDER_CAP_EXCESS_PER_TEU = 5e7

HARD_TIME_WINDOW = False

PEN_MISS_TT            = 5e7
PEN_MISS_ALLOC         = 1e9
PEN_CAP_EXCESS_PER_TEU = 5e7
PEN_LATE_PER_TEU_H = 1.0e6

WAITING_COST_PER_TEU_HOUR_DEFAULT    = 0.8
WAIT_EMISSION_gCO2_per_TEU_H_DEFAULT = 0.0

# ── GA hyper-parameters  [TUNED Stage 2 → G7] ────────────
CROSSOVER_RATE  = 0.90       # ← Stage 2 best (G7: pc=0.90, pm=0.15)
MUTATION_RATE   = 0.15       # ← Stage 2 best

W_ADD  = 0.25
W_DEL  = 0.25
W_MOD  = 0.25
W_MODE = 0.25
OPS    = ["add", "del", "mod", "mode"]

# ── Path library  [TUNED Stage 3] ────────────────────────
PATHS_TOPK_PER_CRITERION = 15        # ← Stage 3 best
PATH_LIB_CAP_TOTAL       = 45        # ← TopK × 3
DFS_MAX_PATHS_PER_OD     = 200       # ← Stage 3 best

CROSSOVER_SEGMENT_PROB = 0.50

# ── Feasibility boost ────────────────────────────────────
MIN_FEASIBLE_SOLUTIONS       = 10
FEASIBLE_BOOST_ROUNDS        = 20
FEASIBLE_BOOST_MUTATION_RATE = 0.60
FEASIBLE_BOOST_TOPK_PARENTS  = 10

# ── Metrics ──────────────────────────────────────────────
HV_EVERY     = 5
HV_SAMPLES   = 50000
METRIC_EVERY = 5

PSTAR_TAIL_GENS   = 30
PSTAR_CAP_PER_GEN = 40
PSTAR_MAX_TOTAL   = 50000

HV_REF_NORM = (1.2, 1.2, 1.2)
HV_MC_SEED  = 12345

DEFAULT_PENALTY_PER_TEU_H = 65.0

NUM_OBJ = 3  # Cost, Emission, Time


# ════════════════════════════════════════════════════════
# Corridor constraint helpers
# ════════════════════════════════════════════════════════

def china_border_monotone_ok(nodes: List[str], node_region: Dict[str, str]) -> bool:
    passed_border = False
    left_china    = False
    start_is_border = (len(nodes) > 0 and nodes[0] in CHINA_BORDER_NODES)
    for i, n in enumerate(nodes):
        r        = str(node_region.get(n, "")).strip()
        in_china  = (r in CHINA_REGIONS)
        is_border = (n in CHINA_BORDER_NODES)
        if left_china and in_china:
            return False
        if in_china:
            if passed_border:
                return False
            if is_border and not (i == 0 and start_is_border):
                passed_border = True
        else:
            left_china = True
    return True


def region_monotone_ok(nodes: List[str], node_region: Dict[str, str]) -> bool:
    max_level = -1
    for n in nodes:
        grp   = NODE_GROUP.get(n, "")
        level = CORRIDOR_ORDER.get(grp, -1)
        if level < 0:
            continue
        if level < max_level:
            return False
        max_level = level
    return True


# ════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════

def normalize_mode(mode_raw: str) -> str:
    m = str(mode_raw).strip().lower()
    if m in {"railway", "rail"}:      return "rail"
    if m in {"road", "truck"}:        return "road"
    if m in {"water", "ship", "sea"}: return "water"
    return m


def safe_float(x, default=0.0) -> float:
    try:
        if pd.isna(x): return default
    except Exception:
        pass
    try:
        return float(x)
    except Exception:
        return default


def parse_distance_km(x) -> float:
    s       = str(x)
    cleaned = "".join(ch for ch in s if (ch.isdigit() or ch == "."))
    return float(cleaned) if cleaned else 0.0


def norm_region(x: str) -> str:
    s  = str(x).strip()
    if not s or s.lower() in {"nan", "none", ""}: return ""
    sl = s.lower()
    if sl in {"china", "prc", "chn"}:                return "CN"
    if sl in {"we", "west europe", "western europe"}: return "WE"
    if sl in {"ee", "east europe", "eastern europe"}: return "EE"
    return s.upper()


def unique_objective_tuples(objs, tol=1e-9):
    out = []
    for o in objs:
        dup = any(all(abs(o[i] - p[i]) <= tol for i in range(NUM_OBJ)) for p in out)
        if not dup:
            out.append(o)
    return out


def _is_bad_text_token(s: str) -> bool:
    if s is None: return True
    t = str(s).strip()
    return t == "" or t.startswith("...")


def _ffill_nan(arr: np.ndarray) -> np.ndarray:
    x = np.array(arr, dtype=float).copy()
    if x.size == 0: return x
    finite_idx = np.where(np.isfinite(x))[0]
    if finite_idx.size == 0: return x
    x[~np.isfinite(x)] = np.nan
    first = finite_idx[0]
    if first > 0:
        x[:first] = x[first]
    for i in range(1, len(x)):
        if np.isnan(x[i]) and np.isfinite(x[i - 1]):
            x[i] = x[i - 1]
    return x


def _finite_points_array(pts):
    if not pts:
        return np.empty((0, NUM_OBJ), dtype=float)
    arr = np.array(pts, dtype=float)
    if arr.ndim != 2 or arr.shape[1] != NUM_OBJ:
        return np.empty((0, NUM_OBJ), dtype=float)
    return arr[np.all(np.isfinite(arr), axis=1)]


# ════════════════════════════════════════════════════════
# Data structures
# ════════════════════════════════════════════════════════

@dataclass
class Arc:
    from_node: str
    to_node: str
    mode: str
    distance: float
    capacity: float
    cost_per_teu_km: float
    emission_per_teu_km: float
    speed_kmh: float
    from_region: str = ""
    to_region:   str = ""


@dataclass
class TimetableEntry:
    from_node: str
    to_node:   str
    mode:      str
    frequency_per_week: float
    first_departure_hour: float
    headway_hours: float


@dataclass
class Batch:
    batch_id:  int
    origin:    str
    destination: str
    quantity:  float
    ET:        float
    LT:        float
    penalty_per_teu_h: float = DEFAULT_PENALTY_PER_TEU_H


@dataclass
class Path:
    path_id:   int
    origin:    str
    destination: str
    nodes:  List[str]
    modes:  List[str]
    arcs:   List[Arc]
    base_cost_per_teu:     float
    base_emission_per_teu: float
    base_travel_time_h:    float

    def __eq__(self, other):
        if not isinstance(other, Path): return NotImplemented
        return self.nodes == other.nodes and self.modes == other.modes

    def __hash__(self):
        return hash((tuple(self.nodes), tuple(self.modes)))


@dataclass
class PathAllocation:
    path:  Path
    share: float

    def __repr__(self):
        chain = ""
        for i, node in enumerate(self.path.nodes[:-1]):
            chain += f"{node}--({self.path.modes[i]})-->"
        chain += self.path.nodes[-1]
        return f"\n    {{ Structure: [{chain}], Share: {self.share:.2%} }}"


# [NSGA-II] Individual carries rank + crowding_distance
@dataclass(eq=False)
class Individual:
    od_allocations: Dict[Tuple[str, str, int], List[PathAllocation]] = field(default_factory=dict)
    objectives:     Tuple[float, float, float] = (float("inf"), float("inf"), float("inf"))
    penalty:        float = 0.0
    feasible:       bool  = False
    feasible_hard:  bool  = False
    vio_breakdown:  Dict[str, float] = field(default_factory=dict)
    rank:               int   = 0           # [NSGA-II] non-dominated front index (0 = best)
    crowding_distance:  float = 0.0         # [NSGA-II] crowding distance


# ════════════════════════════════════════════════════════
# Merge & normalise shares
# ════════════════════════════════════════════════════════

def merge_and_normalize(allocs: List[PathAllocation]) -> List[PathAllocation]:
    if not allocs: return []
    merged: Dict[Path, float] = {}
    for a in allocs:
        merged[a.path] = merged.get(a.path, 0.0) + float(a.share)
    unique_allocs = [PathAllocation(path=p, share=s) for p, s in merged.items()]
    total = sum(a.share for a in unique_allocs)
    if total <= 1e-12:
        avg = 1.0 / max(1, len(unique_allocs))
        for a in unique_allocs: a.share = avg
    else:
        for a in unique_allocs: a.share /= total
    filtered = [a for a in unique_allocs if a.share > 0.05]
    if not filtered:
        best = max(unique_allocs, key=lambda a: a.share)
        best.share = 1.0
        return [best]
    total2 = sum(a.share for a in filtered)
    if abs(total2 - 1.0) > 1e-9:
        for a in filtered: a.share /= total2
    return filtered


# ════════════════════════════════════════════════════════
# Load data
# ════════════════════════════════════════════════════════

def load_carbon_tax_map(xls):
    out = {}
    if "Carbon_Tax" not in xls.sheet_names: return out
    try:
        df = pd.read_excel(xls, "Carbon_Tax")
        rc = next((c for c in ["RegionGroup","Region","region","RegionCode"] if c in df.columns), None)
        tc = next((c for c in ["CT_USD_per_tonCO2","CarbonTax_$_per_tCO2","CarbonTax","CT","Tax"] if c in df.columns), None)
        if rc and tc:
            for _, row in df.iterrows():
                r = str(row.get(rc,"")).strip()
                if r: out[r] = safe_float(row.get(tc), default=0.0)
        print(f"[INFO] Loaded carbon tax for {len(out)} regions.")
    except Exception as e:
        print(f"[WARN] Failed to read Carbon_Tax ({e}).")
    return out

def load_carbon_tax_applicability(xls):
    """读取 Carbon_Tax_Applicability sheet → {(region, mode): 0 or 1}"""
    out = {}
    if "Carbon_Tax_Applicability" not in xls.sheet_names:
        print("[WARN] Carbon_Tax_Applicability sheet 不存在，默认全部 mode 征税。")
        return out
    try:
        df = pd.read_excel(xls, "Carbon_Tax_Applicability")
        rc = next((c for c in ["RegionGroup", "Region", "region"] if c in df.columns), None)
        mc = next((c for c in ["Mode", "mode"] if c in df.columns), None)
        ac = next((c for c in ["Applies", "Theta", "applicability"] if c in df.columns), None)
        if rc and mc and ac:
            for _, row in df.iterrows():
                r = str(row.get(rc, "")).strip()
                m = normalize_mode(row.get(mc, ""))
                a = int(safe_float(row.get(ac), default=1.0))
                if r and m:
                    out[(r, m)] = a
        print(f"[INFO] Loaded carbon tax applicability: {len(out)} entries.")
    except Exception as e:
        print(f"[WARN] Failed to load Carbon_Tax_Applicability ({e}).")
    return out

def load_border_delay_map(xls):
    """[V1 WIRING] Reads PER-MODE border delay columns.

    The expanded data has `BorderDelay_road_h / _rail_h / _water_h`
    (+ `CustomsClearance_h`) and NO single `Mode`/`BorderDelay_h` column,
    so the old single-column loader silently returned {} → delays were
    never applied. This version reads the per-mode columns and falls back
    to the legacy single-column format for backward compatibility.

    Returns {(node, mode): delay_h}. Delay is applied in
    simulate_path_time_capacity ONLY at BREAK_OF_GAUGE_NODES, so extra
    entries here (e.g. for mislabeled EU cities) are harmless.
    """
    out = {}
    if "Node_Border" not in xls.sheet_names: return out
    try:
        nb = pd.read_excel(xls, "Node_Border")
        nc = next((c for c in ["EnglishName","NodeEN"] if c in nb.columns), None)
        if nc is None: return out
        permode = {"road": "BorderDelay_road_h",
                   "rail": "BorderDelay_rail_h",
                   "water": "BorderDelay_water_h"}
        has_permode = any(col in nb.columns for col in permode.values())
        cc_col = "CustomsClearance_h" if "CustomsClearance_h" in nb.columns else None
        if has_permode:
            for _, row in nb.iterrows():
                n = str(row.get(nc, "")).strip()
                if not n: continue
                cc = safe_float(row.get(cc_col), default=0.0) if cc_col else 0.0
                for m, col in permode.items():
                    if col in nb.columns:
                        d = safe_float(row.get(col), default=0.0)
                        if d > 0.0 or cc > 0.0:
                            out[(n, m)] = d + cc
            print(f"[INFO] Loaded border delay map (per-mode +customs): {len(out)} entries.")
        else:
            mc = next((c for c in ["Mode","mode"] if c in nb.columns), None)
            dc = next((c for c in ["BorderDelay_h","Delay_h","BD"] if c in nb.columns), None)
            if nc and mc and dc:
                for _, row in nb.iterrows():
                    n = str(row.get(nc,"")).strip()
                    m = normalize_mode(row.get(mc,""))
                    if n and m:
                        out[(n, m)] = safe_float(row.get(dc), default=0.0)
            print(f"[INFO] Loaded border delay map (legacy): {len(out)} entries.")
    except Exception as e:
        print(f"[WARN] Failed to load border delay ({e}).")
    return out


def load_border_capacity_map(xls):
    """[V1 WIRING] Reads BorderCapacity_TEUday + BackgroundFlow_TEUday.

    These columns existed in the data but were never read by the model.
    Returns (capacity_map, background_map) keyed by node name. Only
    positive values are stored. The capacity soft-constraint in
    evaluate_individual applies only at BOTTLENECK_NODES with cap>0.
    """
    cap, bg = {}, {}
    if "Node_Border" not in xls.sheet_names: return cap, bg
    try:
        nb = pd.read_excel(xls, "Node_Border")
        nc = next((c for c in ["EnglishName","NodeEN"] if c in nb.columns), None)
        capc = "BorderCapacity_TEUday" if "BorderCapacity_TEUday" in nb.columns else None
        bgc  = "BackgroundFlow_TEUday" if "BackgroundFlow_TEUday" in nb.columns else None
        if nc:
            for _, row in nb.iterrows():
                n = str(row.get(nc, "")).strip()
                if not n: continue
                if capc:
                    v = safe_float(row.get(capc), default=0.0)
                    if v > 0.0: cap[n] = v
                if bgc:
                    v = safe_float(row.get(bgc), default=0.0)
                    if v > 0.0: bg[n] = v
        print(f"[INFO] Loaded border capacity: {len(cap)} nodes | "
              f"background flow: {len(bg)} nodes.")
    except Exception as e:
        print(f"[WARN] Failed to load border capacity ({e}).")
    return cap, bg


def load_emission_factor_map(xls):
    out = {}
    if "Emission_Factors" not in xls.sheet_names: return out
    try:
        df  = pd.read_excel(xls, "Emission_Factors")
        mc  = next((c for c in ["Mode","mode"] if c in df.columns), None)
        rc  = next((c for c in ["RegionGroup","Region","region","RegionCode"] if c in df.columns), None)
        efc = next((c for c in ["gCO2_per_TEU_km_assuming10t","EmissionFactor","Emission_gCO2_per_TEU_km","EF","value"] if c in df.columns), None)
        if mc and rc and efc:
            for _, row in df.iterrows():
                m = normalize_mode(row.get(mc,""))
                r = str(row.get(rc,"")).strip()
                if m and r: out[(m, r)] = safe_float(row.get(efc), default=0.0)
        print(f"[INFO] Loaded emission factors for {len(out)} (mode, region) pairs.")
    except Exception as e:
        print(f"[WARN] Failed to read Emission_Factors ({e}).")
    return out


def load_mode_speeds(xls):
    out = {}
    if "Mode_Speeds" not in xls.sheet_names: return out
    try:
        df  = pd.read_excel(xls, "Mode_Speeds")
        mc  = next((c for c in ["Mode","mode"] if c in df.columns), None)
        spc = next((c for c in ["Speed_kmh","speed_kmh","Speed"] if c in df.columns), None)
        if mc and spc:
            for _, row in df.iterrows():
                m = normalize_mode(row.get(mc,""))
                if m: out[m] = safe_float(row.get(spc), default=0.0)
        print(f"[INFO] Loaded mode speeds: {out}")
    except Exception as e:
        print(f"[WARN] Failed to read Mode_Speeds ({e}).")
    return out


def load_transshipment_map(xls):
    out = {}
    if "Transshipment" not in xls.sheet_names: return out
    try:
        df   = pd.read_excel(xls, "Transshipment")
        ndc  = next((c for c in ["Node","NodeEN","EnglishName"] if c in df.columns), None)
        imc  = next((c for c in ["InMode","FromMode","mode_in"] if c in df.columns), None)
        omc  = next((c for c in ["OutMode","ToMode","mode_out"] if c in df.columns), None)
        cstc = next((c for c in ["TransCost","Cost","trans_cost","Cost_per_TEU"] if c in df.columns), None)
        tmc  = next((c for c in ["TransTime_h","Time_h","trans_time_h","Time"] if c in df.columns), None)
        if ndc and imc and omc:
            for _, row in df.iterrows():
                node     = str(row.get(ndc,"")).strip()
                in_mode  = normalize_mode(row.get(imc,""))
                out_mode = normalize_mode(row.get(omc,""))
                if node and in_mode and out_mode:
                    out[(node, in_mode, out_mode)] = {
                        "cost_per_teu": safe_float(row.get(cstc), default=0.0) if cstc else 0.0,
                        "time_h":       safe_float(row.get(tmc),  default=0.0) if tmc  else 0.0,
                    }
        print(f"[INFO] Loaded transshipment entries: {len(out)}")
    except Exception as e:
        print(f"[WARN] Failed to read Transshipment ({e}).")
    return out


def load_waiting_params(xls):
    wc = WAITING_COST_PER_TEU_HOUR_DEFAULT
    we = WAIT_EMISSION_gCO2_per_TEU_H_DEFAULT
    if "Waiting_Costs" not in xls.sheet_names: return wc, we
    try:
        df = pd.read_excel(xls, "Waiting_Costs")
        def pick(colnames, default):
            for c in colnames:
                if c in df.columns:
                    vals = df[c].dropna().tolist()
                    if vals: return safe_float(vals[0], default=default)
            return default
        wc = pick(["WaitingCost_per_TEU_h","WaitCost_per_TEU_h"], wc)
        we = pick(["WaitEmission_gCO2_per_TEU_h","WaitingEmission_gCO2_per_TEU_h"], we)
        print(f"[INFO] Loaded waiting params: cost={wc}, emission={we}")
    except Exception as e:
        print(f"[WARN] Failed to read Waiting_Costs ({e}).")
    return wc, we


def load_network_from_extended(filename: str):
    global CHINA_BORDER_NODES, NODE_GROUP, BORDER_CAPACITY, BACKGROUND_FLOW
    xls = pd.ExcelFile(filename)

    carbon_tax_map      = load_carbon_tax_map(xls)
    emission_factor_map = load_emission_factor_map(xls)
    mode_speeds_map     = load_mode_speeds(xls)
    trans_map           = load_transshipment_map(xls)
    border_delay_map    = load_border_delay_map(xls)
    theta_rm = load_carbon_tax_applicability(xls)

    # [V1 WIRING] populate node capacity & background-flow globals
    BORDER_CAPACITY, BACKGROUND_FLOW = load_border_capacity_map(xls)
    _active_caps = {n: BORDER_CAPACITY[n] for n in BREAK_OF_GAUGE_NODES
                    if n in BORDER_CAPACITY}
    print(f"[INFO] Active break-of-gauge capacities: {_active_caps}")
    print(f"[INFO] Active break-of-gauge background flow: "
          f"{ {n: BACKGROUND_FLOW.get(n,0.0) for n in BREAK_OF_GAUGE_NODES if n in BACKGROUND_FLOW} }")

    nodes_df   = pd.read_excel(xls, "Nodes")
    node_names = nodes_df["EnglishName"].astype(str).str.strip().tolist()

    node_region = {
        str(name).strip(): norm_region(reg)
        for name, reg in zip(nodes_df["EnglishName"], nodes_df["Region"])
    }

    if "RegionGroup" in nodes_df.columns:
        NODE_GROUP = {
            str(name).strip(): str(grp).strip()
            for name, grp in zip(nodes_df["EnglishName"], nodes_df["RegionGroup"])
            if str(grp).strip() not in ("", "nan", "None")
        }

    if "Node_Border" in xls.sheet_names:
        try:
            nb_df = pd.read_excel(xls, "Node_Border")
            loaded_borders = set()
            for _, row in nb_df.iterrows():
                region_val = str(row.get("Region","")).strip()
                is_border  = safe_float(row.get("IsBorderNode", 0), default=0.0) == 1.0
                if region_val == "CN" and is_border:
                    node_name = str(row.get("EnglishName","")).strip()
                    if node_name: loaded_borders.add(node_name)
            if loaded_borders:
                CHINA_BORDER_NODES = loaded_borders
                print(f"[INFO] Loaded China border nodes ({len(CHINA_BORDER_NODES)}): {sorted(CHINA_BORDER_NODES)}")
        except Exception as e:
            print(f"[WARN] Failed to load Node_Border: {e}. Using defaults.")
    else:
        print(f"[INFO] Node_Border sheet not found. Using defaults: {sorted(CHINA_BORDER_NODES)}")

    CHINA_BORDER_NODES.update({"Ningbo", "Shanghai"})

    node_hold_cost:  Dict[str, float] = {}
    node_proc_cost:  Dict[str, float] = {}
    node_trans_cost: Dict[str, float] = {}   # 新增：W_j^trans
    for _, row in nodes_df.iterrows():
        n = str(row.get("EnglishName","")).strip()
        node_hold_cost[n]  = safe_float(row.get("HoldCost_per_TEU_h"),
                                        default=WAITING_COST_PER_TEU_HOUR_DEFAULT)
        node_proc_cost[n]  = safe_float(row.get("ProcCost_per_TEU_h"),
                                        default=0.0)
        node_trans_cost[n] = safe_float(row.get("TransOpCost_per_TEU_h"),
                                        default=node_proc_cost[n])

    SEAPORT_NODES = {"Ningbo", "Shanghai"}
    for n in SEAPORT_NODES:
        if n not in node_region:
            node_region[n] = "CN"
            NODE_GROUP[n]  = "CN"
            if n not in node_names: node_names.append(n)

    waiting_cost_per_teu_h, wait_emis_g_per_teu_h = load_waiting_params(xls)

    arcs_df   = pd.read_excel(xls, "Arcs_All")
    arcs: List[Arc] = []
    cost_cols = ["Cost_$_per_km","Cost_per_km","Cost"]
    emis_cols = ["Emission_gCO2_per_tkm","Emission_gCO2_per_TEU_km","EmissionFactor","Emission"]

    for _, row in arcs_df.iterrows():
        mode  = normalize_mode(row.get("Mode","road"))
        speed = {"road": 75.0, "water": 30.0}.get(mode, 50.0)
        if mode in mode_speeds_map and mode_speeds_map[mode] > 0:
            speed = mode_speeds_map[mode]

        origin = str(row.get("OriginEN","")).strip()
        dest   = str(row.get("DestEN","")).strip()
        if _is_bad_text_token(origin) or _is_bad_text_token(dest): continue

        from_region = str(node_region.get(origin,"")).strip()
        to_region   = str(node_region.get(dest,"")).strip()
        distance    = parse_distance_km(row.get("Distance_km", 0.0))

        if "Capacity_TEUday" in arcs_df.columns and not pd.isna(row.get("Capacity_TEUday", np.nan)):
            capacity = safe_float(row.get("Capacity_TEUday"), default=1e9)
        elif "Capacity_TEUh" in arcs_df.columns and not pd.isna(row.get("Capacity_TEUh", np.nan)):
            capacity = safe_float(row.get("Capacity_TEUh"), default=1e9) * 24.0
        else:
            capacity = 1e9

        cpkm = 0.0
        for c in cost_cols:
            if c in arcs_df.columns:
                val = safe_float(row.get(c), default=None)
                if val is not None and val > 0: cpkm = val; break
        if cpkm <= 1e-9: cpkm = 0.5

        epkm = 0.0
        for c in emis_cols:
            if c in arcs_df.columns:
                epkm = safe_float(row.get(c), default=0.0); break
        if "Emission_gCO2_per_tkm" in arcs_df.columns and epkm > 0:
            epkm = epkm * 10.0
        if (mode, from_region) in emission_factor_map:
            epkm = emission_factor_map[(mode, from_region)]

        arcs.append(Arc(
            from_node=origin, to_node=dest, mode=mode,
            distance=distance, capacity=capacity,
            cost_per_teu_km=cpkm, emission_per_teu_km=epkm,
            speed_kmh=speed, from_region=from_region, to_region=to_region
        ))

    tdf = pd.read_excel(xls, "Timetable")
    timetables: List[TimetableEntry] = []
    for _, row in tdf.iterrows():
        origin    = str(row.get("OriginEN","")).strip()
        dest      = str(row.get("DestEN","")).strip()
        mode_norm = normalize_mode(row.get("Mode",""))
        if _is_bad_text_token(origin) or _is_bad_text_token(dest): continue
        if mode_norm not in {"road","rail","water"}: continue
        freq   = safe_float(row.get("Frequency_per_week"), default=1.0)
        hd_raw = row.get("Headway_Hours", np.nan)
        hd     = 168.0 / max(freq, 1.0) if pd.isna(hd_raw) else safe_float(hd_raw, default=168.0)
        v      = row.get("FirstDepartureHour", np.nan)
        fd     = 0.0
        if not pd.isna(v):
            try:
                s  = str(v).strip()
                fd = float(s.split(":")[0]) if ":" in s else float(s)
            except Exception:
                fd = 0.0
        timetables.append(TimetableEntry(
            from_node=origin, to_node=dest, mode=mode_norm,
            frequency_per_week=freq, first_departure_hour=fd, headway_hours=hd
        ))

    bdf     = pd.read_excel(xls, "Batches")
    bdf     = augment_batches_to_20(bdf, node_region=node_region, random_seed=2026)
    batches: List[Batch] = []
    for _, row in bdf.iterrows():
        origin = str(row.get("OriginEN","")).strip()
        dest   = str(row.get("DestEN","")).strip()
        if node_region.get(origin) in CHINA_REGIONS and node_region.get(dest) in EUROPE_REGIONS:
            batches.append(Batch(
                batch_id=int(row.get("BatchID", 0)),
                origin=origin, destination=dest,
                quantity=safe_float(row.get("QuantityTEU"), default=0.0),
                ET=safe_float(row.get("ET"), default=0.0),
                LT=safe_float(row.get("LT"), default=0.0),
                penalty_per_teu_h=safe_float(row.get("PenaltyCost_per_TEU_h"),
                                             default=DEFAULT_PENALTY_PER_TEU_H)
            ))

    print(f"[INFO] Batches loaded: {len(batches)}")
    return (
        node_names, node_region,
        node_hold_cost, node_proc_cost, node_trans_cost,
        arcs, timetables, batches,
        waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
        carbon_tax_map, emission_factor_map, mode_speeds_map,
        trans_map, border_delay_map, theta_rm
    )


def build_graph(arcs):
    g = {}
    for a in arcs:
        g.setdefault(a.from_node, []).append((a.to_node, a))
    return g


def build_timetable_dict(timetables):
    tt = {}
    for t in timetables:
        tt.setdefault((t.from_node, t.to_node, t.mode), []).append(t)
    return tt


def build_arc_lookup(arcs):
    mp = {}
    for a in arcs:
        k = (a.from_node, a.to_node, a.mode)
        if k not in mp: mp[k] = a
    return mp


# ════════════════════════════════════════════════════════
# Batch augmentation
# ════════════════════════════════════════════════════════

def augment_batches_to_20(bdf, node_region, random_seed=2026):
    df = bdf.copy()
    required_cols = ["BatchID","OriginEN","DestEN","QuantityTEU","ET","LT"]
    if any(c not in df.columns for c in required_cols) or len(df) >= 20:
        return df
    china_nodes  = [n for n, r in node_region.items()
                    if r in CHINA_REGIONS and n not in CHINA_BORDER_NODES]
    europe_nodes = [n for n, r in node_region.items() if r in EUROPE_REGIONS]
    if not china_nodes or not europe_nodes: return df
    q_vals  = pd.to_numeric(df["QuantityTEU"], errors="coerce").dropna()
    q_min   = int(q_vals.min()) if len(q_vals) else 80
    q_max   = int(q_vals.max()) if len(q_vals) else 150
    lt_vals = pd.to_numeric(df["LT"], errors="coerce").dropna()
    lt_vals = lt_vals[lt_vals >= 300]
    lt_min, lt_max = (int(lt_vals.min()), int(lt_vals.max())) if len(lt_vals) else (360, 504)
    pen_col_exists = "PenaltyCost_per_TEU_h" in df.columns
    if pen_col_exists:
        pv = pd.to_numeric(df["PenaltyCost_per_TEU_h"], errors="coerce").dropna()
        pv = pv[(pv >= 1.0) & (pv <= 500.0)]
        pen_min, pen_max = (float(pv.min()), float(pv.max())) if len(pv) else (30.0, 100.0)
    else:
        pen_min, pen_max = 30.0, 100.0
    existing_ids = set(pd.to_numeric(df["BatchID"], errors="coerce").dropna().astype(int).tolist())
    next_id      = max(existing_ids) + 1 if existing_ids else 11
    rng          = np.random.default_rng(random_seed)
    new_rows     = []
    for i in range(20 - len(df)):
        new_rows.append({
            "BatchID":              next_id + i,
            "OriginEN":             str(rng.choice(china_nodes)),
            "DestEN":               str(rng.choice(europe_nodes)),
            "QuantityTEU":          int(rng.integers(q_min, q_max + 1)),
            "ET":                   0,
            "LT":                   int(rng.integers(lt_min, lt_max + 1)),
            "PenaltyCost_per_TEU_h": round(float(rng.uniform(pen_min, pen_max)), 2),
        })
    df_out = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    print(f"[INFO] Batches augmented: {len(df)} -> {len(df_out)}")
    return df_out


# ════════════════════════════════════════════════════════
# Path library
# ════════════════════════════════════════════════════════

def random_dfs_paths(graph, origin, dest, node_region,
                     max_len=12, max_paths=200, timeout_sec=8.0):
    deadline = time.time() + timeout_sec
    paths, found_set = [], set()
    attempts, max_attempts = 0, max_paths * 20
    while len(paths) < max_paths and attempts < max_attempts:
        if time.time() > deadline: break
        attempts += 1
        node, cur_arcs, visited, cur_nodes, ok = origin, [], {origin}, [origin], True
        for _ in range(max_len):
            if node == dest: break
            neighbors = list(graph.get(node, []))
            if not neighbors: ok = False; break
            random.shuffle(neighbors)
            moved = False
            for nxt, arc in neighbors:
                if nxt in visited: continue
                new_nodes = cur_nodes + [nxt]
                if not china_border_monotone_ok(new_nodes, node_region): continue
                if not region_monotone_ok(new_nodes, node_region): continue
                cur_arcs.append(arc); visited.add(nxt)
                cur_nodes.append(nxt); node = nxt; moved = True; break
            if not moved: ok = False; break
        if ok and node == dest and cur_arcs:
            key = (tuple(cur_nodes), tuple(a.mode for a in cur_arcs))
            if key not in found_set:
                found_set.add(key); paths.append(cur_arcs)
    return paths


def repair_arc_seq_with_road_fallback(arc_seq, tt_dict, arc_lookup):
    new_seq = []
    for arc in arc_seq:
        if arc.mode == "road": new_seq.append(arc); continue
        if tt_dict.get((arc.from_node, arc.to_node, arc.mode), []):
            new_seq.append(arc); continue
        k_road = (arc.from_node, arc.to_node, "road")
        if k_road in arc_lookup: new_seq.append(arc_lookup[k_road])
        else: return None
    return new_seq


def select_topk_by_cost_time_emis(paths, k=30, cap_total=90):
    if not paths: return []
    by_cost = sorted(paths, key=lambda p: p.base_cost_per_teu)
    by_time = sorted(paths, key=lambda p: p.base_travel_time_h)
    by_emis = sorted(paths, key=lambda p: p.base_emission_per_teu)
    picked, used = [], set()
    for lst in [by_cost, by_time, by_emis]:
        for p in lst[:k]:
            if p not in used: picked.append(p); used.add(p)
    return picked[:cap_total] if cap_total else picked


def build_path_library(node_names, node_region, arcs, batches, tt_dict, arc_lookup):
    graph    = build_graph(arcs)
    path_lib = {}
    next_pid = 0
    for b in batches:
        od = (b.origin, b.destination)
        if od in path_lib: continue
        arc_paths = random_dfs_paths(graph, b.origin, b.destination,
                                     node_region=node_region, max_len=12,
                                     max_paths=DFS_MAX_PATHS_PER_OD)
        paths_od = []
        for arc_seq in arc_paths:
            repaired = repair_arc_seq_with_road_fallback(arc_seq, tt_dict, arc_lookup)
            if repaired is None: continue
            nodes = [repaired[0].from_node] + [a.to_node for a in repaired]
            if len(set(nodes)) != len(nodes): continue
            if not region_monotone_ok(nodes, node_region): continue
            if not china_border_monotone_ok(nodes, node_region): continue
            modes = [a.mode for a in repaired]
            paths_od.append(Path(
                path_id=next_pid, origin=b.origin, destination=b.destination,
                nodes=nodes, modes=modes, arcs=repaired,
                base_cost_per_teu=sum(a.cost_per_teu_km * a.distance for a in repaired),
                base_emission_per_teu=sum(a.emission_per_teu_km * a.distance for a in repaired),
                base_travel_time_h=sum(a.distance / max(a.speed_kmh, 1.0) for a in repaired),
            ))
            next_pid += 1
        if paths_od:
            path_lib[od] = select_topk_by_cost_time_emis(
                paths_od, k=PATHS_TOPK_PER_CRITERION, cap_total=PATH_LIB_CAP_TOTAL)

    removed = 0
    for od in list(path_lib.keys()):
        before = len(path_lib[od])
        path_lib[od] = [p for p in path_lib[od]
                        if china_border_monotone_ok(p.nodes, node_region)]
        removed += before - len(path_lib[od])
        if not path_lib[od]: del path_lib[od]
    if removed: print(f"[WARN] Post-filter removed {removed} paths.")
    else:       print("[INFO] All paths pass border monotonicity. ✅")
    return path_lib


def sanity_check_path_lib(batches, path_lib):
    missing = [(b.batch_id, (b.origin, b.destination))
               for b in batches if not path_lib.get((b.origin, b.destination), [])]
    if missing:
        for bid, od in missing[:20]:
            print(f"[SANITY] ❌ missing paths Batch {bid} OD={od}")
        raise RuntimeError("Path library missing some ODs.")
    print("[SANITY] ✅ All batches have paths.")


def repair_missing_allocations(ind, batches, path_lib):
    for b in batches:
        key = (b.origin, b.destination, b.batch_id)
        if ind.od_allocations.get(key, []): continue
        paths = path_lib.get((b.origin, b.destination), [])
        if paths:
            ind.od_allocations[key] = [PathAllocation(path=paths[0], share=1.0)]


# ════════════════════════════════════════════════════════
# Simulation & evaluation
# ════════════════════════════════════════════════════════

def next_departure_time_programB(t: float, entries: List[TimetableEntry]) -> float:
    best = float("inf")
    for e in entries:
        if t <= e.first_departure_hour:
            dep = e.first_departure_hour
        else:
            waited = t - e.first_departure_hour
            n      = math.ceil(waited / max(e.headway_hours, 1e-6))
            dep    = e.first_departure_hour + n * e.headway_hours
        if dep < best: best = dep
    return best if best < float("inf") else t


def simulate_path_time_capacity(
    path: Path, batch: Batch, flow_teu: float,
    tt_dict: Dict, arc_flow_map: Dict,
    trans_map: Optional[Dict] = None,
    border_delay_map: Optional[Dict] = None,
    node_flow_map: Optional[Dict] = None,   # [V1 WIRING] per-(node,day) throughput
) -> Tuple[float, List[Tuple[str, float, float]], int]:
    t                = float(batch.ET)
    miss_tt          = 0
    trans_map        = trans_map or {}
    border_delay_map = border_delay_map or {}
    prev_arc         = None
    node_wait_list: List[Tuple[str, float, float]] = []

    for arc in path.arcs:
        cur_node       = arc.from_node
        arc_trans_wait = 0.0

        if prev_arc is not None and prev_arc.mode != arc.mode:
            rec = trans_map.get((cur_node, prev_arc.mode, arc.mode))
            if rec:
                th = safe_float(rec.get("time_h"), default=0.0)
                if th > 0: t += th; arc_trans_wait += th

        # [V1 WIRING] fixed border delay now applies at ALL break-of-gauge
        # nodes (incl. non-CN: Dostyk, Altynkol, Zabaykalsk, Brest,
        # Malaszewicze), not just China. CHINA_BORDER_NODES is left for
        # routing monotonicity only.
        if cur_node in BREAK_OF_GAUGE_NODES:
            bd = border_delay_map.get((cur_node, arc.mode), 0.0)
            if bd > 0: t += bd; arc_trans_wait += bd

        travel_arc = arc.distance / max(arc.speed_kmh, 1.0)
        entries    = [] if arc.mode == "road" else \
                     tt_dict.get((cur_node, arc.to_node, arc.mode), [])
        if arc.mode != "road" and not entries:
            miss_tt += 1; return float("inf"), [], miss_tt

        dep            = t if not entries else next_departure_time_programB(t, entries)
        arc_sched_wait = max(0.0, dep - t)
        node_wait_list.append((cur_node, arc_sched_wait, arc_trans_wait))

        arr  = dep + travel_arc
        slot = int(dep // 24)
        akey = (cur_node, arc.to_node, arc.mode)
        arc_flow_map[(akey, slot)] = arc_flow_map.get((akey, slot), 0.0) + flow_teu

        # [V1 WIRING] accumulate departing throughput at bottleneck nodes
        # for per-day node capacity check (counted once per transit).
        if node_flow_map is not None and cur_node in BOTTLENECK_NODES:
            node_flow_map[(cur_node, slot)] = \
                node_flow_map.get((cur_node, slot), 0.0) + flow_teu

        t        = arr
        prev_arc = arc

    return (t - batch.ET), node_wait_list, miss_tt


def evaluate_individual(
    ind, batches, arcs, tt_dict,
    waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
    node_hold_cost=None, node_proc_cost=None,
    carbon_tax_map=None, trans_map=None, border_delay_map=None,
    theta_rm=None, node_trans_cost=None,
):
    node_hold_cost   = node_hold_cost   or {}
    node_proc_cost   = node_proc_cost   or {}
    carbon_tax_map   = carbon_tax_map   or {}
    trans_map        = trans_map        or {}
    border_delay_map = border_delay_map or {}
    theta_rm         = theta_rm         or {}
    node_trans_cost  = node_trans_cost  or {}

    total_cost = total_emission_g = makespan = 0.0
    arc_flow_map: Dict = {}
    node_flow_map: Dict = {}   # [V1 WIRING] (node, day-slot) -> TEU throughput
    arc_caps = {(a.from_node, a.to_node, a.mode): a.capacity for a in arcs}

    miss_alloc = miss_tt = 0
    cap_excess = late_teu_h_total = wait_teu_h_total = 0.0
    trans_teu_h_total = trans_cost_total = carbon_cost_total = 0.0

    for b in batches:
        key    = (b.origin, b.destination, b.batch_id)
        allocs = ind.od_allocations.get(key, [])
        if not allocs: miss_alloc += 1; continue

        batch_finish = b.ET
        y_jmn_k: Dict[Tuple[str, str, str], float] = {}
        for alloc in allocs:
            if alloc.share <= 1e-12: continue
            flow = alloc.share * b.quantity
            p    = alloc.path

            travel_time, node_wait_list, mtt = simulate_path_time_capacity(
                p, b, flow, tt_dict, arc_flow_map,
                trans_map=trans_map, border_delay_map=border_delay_map,
                node_flow_map=node_flow_map)
            if math.isinf(travel_time): miss_tt += mtt; continue

            total_cost       += p.base_cost_per_teu * flow
            total_emission_g += p.base_emission_per_teu * flow

            # ── 聚合 y_jmn^k：在节点 j 从 mode m_in 转到 mode m_out 的 TEU ──
            for i in range(len(p.arcs) - 1):
                if p.arcs[i].mode != p.arcs[i+1].mode:
                    j_node = p.arcs[i+1].from_node
                    m_in   = p.arcs[i].mode
                    m_out  = p.arcs[i+1].mode
                    y_jmn_k[(j_node, m_in, m_out)] = \
                        y_jmn_k.get((j_node, m_in, m_out), 0.0) + flow

            cc = 0.0
            for arc in p.arcs:
                region = getattr(arc, "from_region", "")
                # Θ_{r,m}: 缺失时默认 1（向后兼容，全征税）
                theta = theta_rm.get((region, arc.mode), 1)
                if theta == 0:
                    continue
                tax_rate  = float(carbon_tax_map.get(region, 0.0))
                emis_tons = arc.emission_per_teu_km * arc.distance * flow / 1e6
                cc += emis_tons * tax_rate
            total_cost += cc; carbon_cost_total += cc


            for (wnode, sched_h, trans_h) in node_wait_list:
                hold_rate = node_hold_cost.get(wnode, WAITING_COST_PER_TEU_HOUR_DEFAULT)
                proc_rate = node_proc_cost.get(wnode, 0.0)
                if sched_h > 0.0:
                    total_cost       += hold_rate * flow * sched_h
                    total_emission_g += wait_emis_g_per_teu_h * flow * sched_h
                    wait_teu_h_total += flow * sched_h
                if trans_h > 0.0:
                    total_cost        += proc_rate * flow * trans_h
                    trans_teu_h_total += flow * trans_h

            arrival_time = b.ET + travel_time
            batch_finish = max(batch_finish, arrival_time)
            if arrival_time > b.LT:
                late_h            = flow * (arrival_time - b.LT)
                late_teu_h_total += late_h
                total_cost       += b.penalty_per_teu_h * late_h
        # ── 用聚合后的 y_jmn^k 统一计算转运成本 ──
        for (j_node, m_in, m_out), y_val in y_jmn_k.items():
            rec       = trans_map.get((j_node, m_in, m_out), {})
            tc_unit   = safe_float(rec.get("cost_per_teu"), default=0.0)
            tt_h      = safe_float(rec.get("time_h"),       default=0.0)
            # (a) 转运处理成本 TC_jmn × y
            total_cost       += tc_unit * y_val
            trans_cost_total += tc_unit * y_val
            # (b) 转运操作成本 W_j^trans × y × TT_jmn （模型公式 12）
            w_trans            = node_trans_cost.get(j_node, 0.0)
            total_cost        += w_trans * y_val * tt_h
            trans_teu_h_total += y_val * tt_h

        makespan = max(makespan, batch_finish)

    for (akey, slot), sf in arc_flow_map.items():
        cap = arc_caps.get(akey, 1e9)
        if sf > cap: cap_excess += (sf - cap)

    # ── [V1 WIRING] node/border daily capacity (soft) + utilisation ──
    # load_j,slot = batch flow through j on that day + daily background flow.
    # u_j = load / BorderCapacity_j ; excess penalised like arc capacity.
    # Applied only at BOTTLENECK_NODES that have a positive capacity.
    border_cap_excess = 0.0
    max_border_util   = 0.0
    border_node_flow: Dict[str, float] = {}   # total batch TEU through node
    border_node_util: Dict[str, float] = {}   # PEAK daily utilisation
    for (nname, slot), nf in node_flow_map.items():
        border_node_flow[nname] = border_node_flow.get(nname, 0.0) + nf
        cap = BORDER_CAPACITY.get(nname, 0.0)
        if cap <= 0.0:
            continue
        load = nf + BACKGROUND_FLOW.get(nname, 0.0)
        u    = load / cap
        if u > border_node_util.get(nname, 0.0):
            border_node_util[nname] = u
        if u > max_border_util:
            max_border_util = u
        if load > cap:
            border_cap_excess += (load - cap)

    penalty = (PEN_MISS_ALLOC * float(miss_alloc) +
               PEN_MISS_TT    * float(miss_tt)    +
               PEN_CAP_EXCESS_PER_TEU * float(cap_excess) +
               PEN_BORDER_CAP_EXCESS_PER_TEU * float(border_cap_excess) +
               PEN_LATE_PER_TEU_H * float(late_teu_h_total))

    ind.objectives    = (float(total_cost), float(total_emission_g), float(makespan))
    ind.penalty       = float(penalty)
    hard_ok = (miss_alloc == 0 and miss_tt == 0 and cap_excess <= 1e-9
               and border_cap_excess <= 1e-9
               and late_teu_h_total <= 1e-9)
    ind.feasible_hard = bool(hard_ok)
    ind.feasible      = bool(hard_ok)
    ind.vio_breakdown = {
        "miss_alloc":  float(miss_alloc),
        "miss_tt":     float(miss_tt),
        "cap_excess":  float(cap_excess),
        "border_cap_excess": float(border_cap_excess),
        "max_border_util":   float(max_border_util),
        "late_teu_h":  float(late_teu_h_total),
        "wait_teu_h":  float(wait_teu_h_total),
        "trans_teu_h": float(trans_teu_h_total),
        "trans_cost":  float(trans_cost_total),
        "carbon_cost": float(carbon_cost_total),
    }
    # [V1 WIRING] per-node detail for sensitivity reporting (not used by GA)
    ind.border_flow = border_node_flow
    ind.border_util = border_node_util


# ════════════════════════════════════════════════════════
# GA operators
# ════════════════════════════════════════════════════════

def clone_gene(alloc):
    return PathAllocation(path=alloc.path, share=float(alloc.share))


def crossover_structural(ind1, ind2, batches):
    child1, child2 = Individual(), Individual()
    for b in batches:
        key = (b.origin, b.destination, b.batch_id)
        g1  = ind1.od_allocations.get(key, [])
        g2  = ind2.od_allocations.get(key, [])
        if not g1 and not g2: continue
        if not g1:
            child1.od_allocations[key] = [clone_gene(x) for x in g2]
            child2.od_allocations[key] = [clone_gene(x) for x in g2]; continue
        if not g2:
            child1.od_allocations[key] = [clone_gene(x) for x in g1]
            child2.od_allocations[key] = [clone_gene(x) for x in g1]; continue
        cut1, cut2 = random.randint(0, len(g1)), random.randint(0, len(g2))
        c1 = [clone_gene(x) for x in g1[:cut1]] + [clone_gene(x) for x in g2[cut2:]]
        c2 = [clone_gene(x) for x in g2[:cut2]] + [clone_gene(x) for x in g1[cut1:]]
        child1.od_allocations[key] = merge_and_normalize(c1)
        child2.od_allocations[key] = merge_and_normalize(c2)
    return child1, child2


def path_from_arcs(new_arcs, origin, destination, path_id=-1, node_region=None):
    if not new_arcs: return None
    nodes = [new_arcs[0].from_node] + [a.to_node for a in new_arcs]
    if nodes[0] != origin or nodes[-1] != destination: return None
    if len(set(nodes)) != len(nodes): return None
    if node_region is not None:
        if not china_border_monotone_ok(nodes, node_region): return None
        if not region_monotone_ok(nodes, node_region):       return None
    return Path(
        path_id=path_id, origin=origin, destination=destination,
        nodes=nodes, modes=[a.mode for a in new_arcs], arcs=new_arcs,
        base_cost_per_teu=sum(a.cost_per_teu_km * a.distance for a in new_arcs),
        base_emission_per_teu=sum(a.emission_per_teu_km * a.distance for a in new_arcs),
        base_travel_time_h=sum(a.distance / max(a.speed_kmh, 1.0) for a in new_arcs),
    )


def rebuild_path_from_nodes_modes(origin, destination, nodes, modes,
                                   tt_dict, arc_lookup, allow_road_fallback=True):
    if not nodes or len(nodes) < 2 or nodes[0] != origin or nodes[-1] != destination:
        return None
    if len(modes) != len(nodes) - 1 or len(set(nodes)) != len(nodes): return None
    new_arcs = []
    for i in range(len(modes)):
        u, v, m = nodes[i], nodes[i+1], modes[i]
        k = (u, v, m)
        if k not in arc_lookup: return None
        arc = arc_lookup[k]
        if arc.mode != "road" and not tt_dict.get((u, v, arc.mode), []):
            if allow_road_fallback and (u, v, "road") in arc_lookup:
                arc = arc_lookup[(u, v, "road")]
            else: return None
        new_arcs.append(arc)
    return path_from_arcs(new_arcs, origin, destination)


def find_common_internal_nodes(p1, p2):
    return list(set(p1.nodes[1:-1]) & set(p2.nodes[1:-1]))


def perform_single_point_crossover_paths(pA, pB, join_node, tt_dict, arc_lookup):
    if join_node not in pA.nodes or join_node not in pB.nodes: return None
    ia, ib = pA.nodes.index(join_node), pB.nodes.index(join_node)
    return rebuild_path_from_nodes_modes(
        pA.origin, pA.destination,
        pA.nodes[:ia+1] + pB.nodes[ib+1:],
        pA.modes[:ia]   + pB.modes[ib:],
        tt_dict, arc_lookup)


def crossover_common_node(ind1, ind2, batches, tt_dict, arc_lookup):
    child1, child2 = Individual(), Individual()
    for b in batches:
        key = (b.origin, b.destination, b.batch_id)
        g1  = ind1.od_allocations.get(key, [])
        g2  = ind2.od_allocations.get(key, [])
        if not g1 and not g2: continue
        if not g1:
            child1.od_allocations[key] = [clone_gene(x) for x in g2]
            child2.od_allocations[key] = [clone_gene(x) for x in g2]; continue
        if not g2:
            child1.od_allocations[key] = [clone_gene(x) for x in g1]
            child2.od_allocations[key] = [clone_gene(x) for x in g1]; continue
        c1_allocs = [clone_gene(x) for x in g1]
        c2_allocs = [clone_gene(x) for x in g2]
        p1, p2    = random.choice(g1).path, random.choice(g2).path
        common    = find_common_internal_nodes(p1, p2)
        if common:
            join = random.choice(common)
            np1  = perform_single_point_crossover_paths(p1, p2, join, tt_dict, arc_lookup)
            np2  = perform_single_point_crossover_paths(p2, p1, join, tt_dict, arc_lookup)
            if np1: c1_allocs.append(PathAllocation(path=np1, share=0.20))
            if np2: c2_allocs.append(PathAllocation(path=np2, share=0.20))
        child1.od_allocations[key] = merge_and_normalize(c1_allocs)
        child2.od_allocations[key] = merge_and_normalize(c2_allocs)
    return child1, child2


def crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup):
    if random.random() < CROSSOVER_SEGMENT_PROB:
        c1, c2 = crossover_common_node(p1, p2, batches, tt_dict, arc_lookup)
        if c1.od_allocations or c2.od_allocations: return c1, c2
    return crossover_structural(p1, p2, batches)


def random_initial_individual(batches, path_lib, max_paths=3):
    ind = Individual()
    for b in batches:
        paths = path_lib.get((b.origin, b.destination), [])
        if not paths: continue
        k      = random.randint(1, min(max_paths, len(paths)))
        chosen = random.sample(paths, k)
        raw    = [PathAllocation(path=p, share=random.random()) for p in chosen]
        ind.od_allocations[(b.origin, b.destination, b.batch_id)] = merge_and_normalize(raw)
    return ind


def greedy_initial_individual(batches, path_lib):
    ind = Individual()
    for b in batches:
        paths = path_lib.get((b.origin, b.destination), [])
        if not paths: continue
        best  = min(paths, key=lambda p: p.base_travel_time_h)
        ind.od_allocations[(b.origin, b.destination, b.batch_id)] = \
            [PathAllocation(path=best, share=1.0)]
    return ind

def greedy_initial_min_cost(batches, path_lib):
    ind = Individual()
    for b in batches:
        paths = path_lib.get((b.origin, b.destination), [])
        if not paths: continue
        best  = min(paths, key=lambda p: p.base_cost_per_teu)
        ind.od_allocations[(b.origin, b.destination, b.batch_id)] = \
            [PathAllocation(path=best, share=1.0)]
    return ind


def greedy_initial_min_emission(batches, path_lib):
    ind = Individual()
    for b in batches:
        paths = path_lib.get((b.origin, b.destination), [])
        if not paths: continue
        best  = min(paths, key=lambda p: p.base_emission_per_teu)
        ind.od_allocations[(b.origin, b.destination, b.batch_id)] = \
            [PathAllocation(path=best, share=1.0)]
    return ind

def mutate_add(ind, batch, path_lib):
    key    = (batch.origin, batch.destination, batch.batch_id)
    allocs = ind.od_allocations.get(key, [])
    pool   = path_lib.get((batch.origin, batch.destination), [])
    if not pool: return False
    cur   = {a.path for a in allocs}
    cands = [p for p in pool if p not in cur]
    if not cands:
        if allocs:
            allocs[random.randrange(len(allocs))] = PathAllocation(
                path=random.choice(pool), share=0.2)
            ind.od_allocations[key] = merge_and_normalize(allocs)
            return True
        return False
    allocs.append(PathAllocation(path=random.choice(cands), share=0.2))
    ind.od_allocations[key] = merge_and_normalize(allocs)
    return True


def mutate_del(ind, batch):
    key    = (batch.origin, batch.destination, batch.batch_id)
    allocs = ind.od_allocations.get(key, [])
    if len(allocs) <= 1: return False
    allocs.pop(random.randrange(len(allocs)))
    ind.od_allocations[key] = merge_and_normalize(allocs)
    return True


def mutate_mod(ind, batch):
    key    = (batch.origin, batch.destination, batch.batch_id)
    allocs = ind.od_allocations.get(key, [])
    if not allocs: return False
    random.choice(allocs).share *= random.uniform(0.5, 1.5)
    ind.od_allocations[key] = merge_and_normalize(allocs)
    return True


def mutate_mode(ind, batch, tt_dict, arc_lookup, max_trials=20):
    key    = (batch.origin, batch.destination, batch.batch_id)
    allocs = ind.od_allocations.get(key, [])
    if not allocs: return False
    idx       = random.randrange(len(allocs))
    old_alloc = allocs[idx]
    p         = old_alloc.path
    if not p.arcs: return False
    arc_i     = random.randrange(len(p.arcs))
    old_arc   = p.arcs[arc_i]
    u, v      = old_arc.from_node, old_arc.to_node
    for _ in range(max_trials):
        new_mode = random.choice([m for m in ["road","rail","water"] if m != old_arc.mode])
        k_arc    = (u, v, new_mode)
        if k_arc not in arc_lookup: continue
        if new_mode != "road" and not tt_dict.get((u, v, new_mode), []): continue
        new_arcs       = list(p.arcs)
        new_arcs[arc_i] = arc_lookup[k_arc]
        new_path        = path_from_arcs(new_arcs, p.origin, p.destination)
        if new_path is None: continue
        allocs_new      = deepcopy(allocs)
        allocs_new[idx] = PathAllocation(path=new_path, share=old_alloc.share)
        ind.od_allocations[key] = merge_and_normalize(allocs_new)
        return True
    return False


_FIXED_OP_WEIGHTS = [W_ADD, W_DEL, W_MOD, W_MODE]
_FIXED_OP_TOTAL   = sum(_FIXED_OP_WEIGHTS)
_FIXED_OP_PROBS   = [w / _FIXED_OP_TOTAL for w in _FIXED_OP_WEIGHTS]


def sample_operator() -> str:
    r, cum = random.random(), 0.0
    for op, prob in zip(OPS, _FIXED_OP_PROBS):
        cum += prob
        if r <= cum: return op
    return OPS[-1]


def apply_mutation_op(ind, op, batch, path_lib, tt_dict, arc_lookup):
    if op == "add":  return mutate_add(ind, batch, path_lib)
    if op == "del":  return mutate_del(ind, batch)
    if op == "mod":  return mutate_mod(ind, batch)
    if op == "mode": return mutate_mode(ind, batch, tt_dict, arc_lookup)
    return False


def mutate_fixed(
    ind, batches, path_lib, tt_dict, arc_lookup,
    arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
    node_hold_cost=None, node_proc_cost=None,
    carbon_tax_map=None, trans_map=None, border_delay_map=None,
    theta_rm=None, node_trans_cost=None,
):
    batch = random.choice(batches)
    op    = sample_operator()
    ok    = apply_mutation_op(ind, op, batch, path_lib, tt_dict, arc_lookup)
    if not ok:
        return op, False
    repair_missing_allocations(ind, batches, path_lib)
    evaluate_individual(
        ind, batches, arcs, tt_dict,
        waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
        node_hold_cost=node_hold_cost, node_proc_cost=node_proc_cost,
        carbon_tax_map=carbon_tax_map, trans_map=trans_map,
        border_delay_map=border_delay_map,
        theta_rm=theta_rm, node_trans_cost=node_trans_cost,
    )
    return op, True


# ════════════════════════════════════════════════════════
# Dominance helper
# ════════════════════════════════════════════════════════

def dominates(a, b):
    if a.feasible and not b.feasible: return True
    if b.feasible and not a.feasible: return False
    if a.feasible and b.feasible:
        return (all(x <= y for x, y in zip(a.objectives, b.objectives)) and
                any(x <  y for x, y in zip(a.objectives, b.objectives)))
    if a.penalty < b.penalty - 1e-12: return True
    if b.penalty < a.penalty - 1e-12: return False
    return (all(x <= y for x, y in zip(a.objectives, b.objectives)) and
            any(x <  y for x, y in zip(a.objectives, b.objectives)))


def unique_individuals_by_objectives(front, tol=1e-3):
    uniq, seen = [], []
    for ind in front:
        obj = ind.objectives
        if not any(all(abs(obj[i]-o[i]) <= tol for i in range(NUM_OBJ)) for o in seen):
            seen.append(obj); uniq.append(ind)
    return uniq


# ════════════════════════════════════════════════════════
# NSGA-II core components
# ════════════════════════════════════════════════════════

def fast_non_dominated_sort(pop: List[Individual]) -> List[List[int]]:
    """
    Deb et al. (2002) fast non-dominated sort.
    Returns fronts: list of index lists. fronts[0] = Pareto-front.
    Also writes ind.rank in-place (0-indexed).
    """
    N = len(pop)
    S         = [[] for _ in range(N)]
    n_dom     = [0]  * N
    rank      = [0]  * N
    fronts: List[List[int]] = [[]]

    for p in range(N):
        for q in range(N):
            if p == q: continue
            if dominates(pop[p], pop[q]):
                S[p].append(q)
            elif dominates(pop[q], pop[p]):
                n_dom[p] += 1
        if n_dom[p] == 0:
            rank[p] = 0
            fronts[0].append(p)

    i = 0
    while fronts[i]:
        nxt: List[int] = []
        for p in fronts[i]:
            for q in S[p]:
                n_dom[q] -= 1
                if n_dom[q] == 0:
                    rank[q] = i + 1
                    nxt.append(q)
        i += 1
        fronts.append(nxt)
    fronts.pop()  # drop trailing empty front

    for idx, r in enumerate(rank):
        pop[idx].rank = int(r)
    return fronts


def crowding_distance_assignment(pop: List[Individual],
                                  indices: List[int]) -> None:
    """Assign crowding distance to individuals in the given front (in place)."""
    if not indices:
        return
    L = len(indices)
    for idx in indices:
        pop[idx].crowding_distance = 0.0
    if L <= 2:
        for idx in indices:
            pop[idx].crowding_distance = float("inf")
        return

    for m in range(NUM_OBJ):
        sorted_idx = sorted(indices, key=lambda i: pop[i].objectives[m])
        f_min = pop[sorted_idx[0]].objectives[m]
        f_max = pop[sorted_idx[-1]].objectives[m]
        rng   = f_max - f_min
        pop[sorted_idx[0]].crowding_distance  = float("inf")
        pop[sorted_idx[-1]].crowding_distance = float("inf")
        if rng <= 1e-12:
            continue
        for k in range(1, L - 1):
            prev_o = pop[sorted_idx[k - 1]].objectives[m]
            next_o = pop[sorted_idx[k + 1]].objectives[m]
            pop[sorted_idx[k]].crowding_distance += (next_o - prev_o) / rng


def nsga2_environmental_selection(combined: List[Individual],
                                   pop_size: int) -> List[Individual]:
    """
    1. Non-dominated sort R = P ∪ Q
    2. Fill new population by fronts F1, F2, ... until next front overflows
    3. Splitting front: keep highest crowding distance
    """
    fronts = fast_non_dominated_sort(combined)
    new_pop: List[Individual] = []
    for front in fronts:
        if not front:
            continue
        crowding_distance_assignment(combined, front)
        if len(new_pop) + len(front) <= pop_size:
            new_pop.extend(combined[i] for i in front)
            if len(new_pop) == pop_size:
                break
        else:
            remaining = pop_size - len(new_pop)
            front_sorted = sorted(
                front, key=lambda i: combined[i].crowding_distance, reverse=True)
            new_pop.extend(combined[i] for i in front_sorted[:remaining])
            break
    return new_pop[:pop_size]


def nsga2_binary_tournament(pop: List[Individual]) -> Individual:
    """Crowded-comparison tournament: lower rank wins; tie → larger crowding dist wins."""
    a, b = random.sample(pop, 2)
    if a.rank < b.rank: return a
    if b.rank < a.rank: return b
    return a if a.crowding_distance >= b.crowding_distance else b


# ════════════════════════════════════════════════════════
# Feasibility boost  (uses rank instead of SPEA2 fitness)
# ════════════════════════════════════════════════════════

def _select_boost_parents(pop, topk=FEASIBLE_BOOST_TOPK_PARENTS):
    feasible   = sorted([i for i in pop if i.feasible],
                        key=lambda x: (x.rank, -x.crowding_distance))
    if len(feasible) >= topk: return feasible[:topk]
    infeasible = sorted([i for i in pop if not i.feasible],
                        key=lambda x: x.penalty)
    return (feasible + infeasible)[:topk]


def feasibility_boost(
    population, batches, path_lib, tt_dict, arc_lookup,
    arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h, eval_kwargs,
    boost_rounds=FEASIBLE_BOOST_ROUNDS,
    boost_mutation_rate=FEASIBLE_BOOST_MUTATION_RATE,
    topk_parents=FEASIBLE_BOOST_TOPK_PARENTS,
):
    parents  = _select_boost_parents(population, topk=topk_parents)
    new_inds = []
    if len(parents) < 2:
        greedy_funcs = [greedy_initial_min_cost,
                        greedy_initial_individual,
                        greedy_initial_min_emission]
        for r in range(boost_rounds):
            ind = greedy_funcs[r % 3](batches, path_lib)
            repair_missing_allocations(ind, batches, path_lib)
            evaluate_individual(ind, batches, arcs, tt_dict,
                                waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
            new_inds.append(ind)
    else:
        for _ in range(boost_rounds):
            p1, p2 = random.sample(parents, 2)
            c1, c2 = crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup)
            for child in (c1, c2):
                repair_missing_allocations(child, batches, path_lib)
                if random.random() < boost_mutation_rate:
                    mutate_fixed(
                        child, batches, path_lib, tt_dict, arc_lookup,
                        arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
                # [V1 FIX] ALWAYS evaluate boost children. Previously, a child
                # that wasn't mutated kept the dataclass default penalty=0.0 /
                # feasible=False, polluting selection with fake "0-penalty"
                # individuals and stalling convergence whenever boost fired.
                evaluate_individual(
                    child, batches, arcs, tt_dict,
                    waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
                new_inds.append(child)
            if len(new_inds) >= boost_rounds: break

    pop_id_idx     = {id(ind): idx for idx, ind in enumerate(population)}
    infeas_sorted  = sorted([i for i in population if not i.feasible],
                            key=lambda x: x.penalty, reverse=True)
    new_sorted     = sorted(new_inds, key=lambda x: (0 if x.feasible else 1, x.penalty))
    num_new_feas   = 0
    replaced_ids: set = set()

    for new_ind in new_sorted:
        for target in infeas_sorted:
            if id(target) in replaced_ids: continue
            if new_ind.feasible or new_ind.penalty < target.penalty:
                idx = pop_id_idx.get(id(target))
                if idx is not None:
                    population[idx] = new_ind
                    replaced_ids.add(id(target))
                    if new_ind.feasible: num_new_feas += 1
                break
    return population, num_new_feas


# ════════════════════════════════════════════════════════
# HV calculator (Monte Carlo, 3-objective)
# ════════════════════════════════════════════════════════

class HypervolumeCalculator:
    def __init__(self, ref_point, num_samples=2000, seed=None):
        self.ref_point   = np.array(ref_point, dtype=float)
        self.num_samples = int(num_samples)
        rng = np.random.default_rng(seed)
        self.samples = rng.uniform(low=0.0, high=self.ref_point,
                                   size=(self.num_samples, NUM_OBJ))

    def calculate_points(self, points):
        if not points: return 0.0
        front = np.array(points, dtype=float)
        valid = front[np.all(front <= self.ref_point, axis=1)]
        if len(valid) == 0: return 0.0
        S   = self.samples[:, np.newaxis, :]
        O   = valid[np.newaxis, :, :]
        dom = np.any(np.all(O <= S, axis=2), axis=1)
        return float(np.sum(dom) / self.num_samples)


# ════════════════════════════════════════════════════════
# Metrics: P*, IGD+, Spacing
# ════════════════════════════════════════════════════════

def dominates_obj(a, b):
    return all(a[i] <= b[i] for i in range(NUM_OBJ)) and any(a[i] < b[i] for i in range(NUM_OBJ))


def nondominated_set(points):
    pts = unique_objective_tuples(points, tol=1e-9)
    return [p for i, p in enumerate(pts)
            if not any(dominates_obj(q, p) for j, q in enumerate(pts) if i != j)]


def normalize_points(points, mins, maxs):
    out = []
    for p in points:
        pp = []
        for i in range(NUM_OBJ):
            rng = maxs[i] - mins[i]
            pp.append(0.0 if rng <= 1e-12 else (p[i]-mins[i])/rng)
        out.append(tuple(pp))
    return out


def clip_points(points, ref):
    return [tuple(min(max(p[i], 0.0), ref[i]) for i in range(NUM_OBJ)) for p in points]


def igd_plus(P_star, A):
    if not P_star or not A: return float("inf")
    P, Q = np.array(P_star, dtype=float), np.array(A, dtype=float)
    return float(np.mean([
        float(np.min(np.sqrt(np.sum(np.maximum(Q - p, 0.0)**2, axis=1))))
        for p in P]))


def spacing_metric(A):
    if not A or len(A) < 2: return 0.0
    Q = np.array(A, dtype=float)
    n = Q.shape[0]
    dmin = []
    for i in range(n):
        diff = Q - Q[i]
        d    = np.sqrt(np.sum(diff**2, axis=1))
        d[i] = np.inf
        dmin.append(float(np.min(d)))
    dmin = np.array(dmin)
    return float(np.sqrt(np.sum((dmin - np.mean(dmin))**2) / max(1, n-1)))


def build_P_star_fast(run_front_hist, tail_gens=PSTAR_TAIL_GENS,
                      cap_per_gen=PSTAR_CAP_PER_GEN, max_total=PSTAR_MAX_TOTAL):
    pts = []
    for hist in run_front_hist:
        tail = hist[-tail_gens:] if tail_gens > 0 else hist
        for gen_front in tail:
            pts.extend(gen_front[:cap_per_gen])
            if len(pts) >= max_total: break
        if len(pts) >= max_total: break
    return nondominated_set(pts)


# ════════════════════════════════════════════════════════
# Export helpers
# ════════════════════════════════════════════════════════

def save_pareto_solutions(pareto, batches, filename="result.txt"):
    pareto = unique_individuals_by_objectives([i for i in pareto if i.feasible], tol=1e-3)
    with open(filename, "w", encoding="utf-8") as f:
        f.write("===== NSGA-II Tuned Baseline Pareto Solutions (3-Objective) =====\n\n")
        if not pareto:
            f.write("NO FEASIBLE SOLUTION FOUND.\n"); return
        for i, ind in enumerate(pareto):
            c, e, t = ind.objectives
            f.write(f"===== Pareto Sol {i} =====\n")
            f.write(f"Cost={c:.6f}  Emission_gCO2={e:.6f}  Time={t:.6f}\n")
            f.write(f"Penalty={ind.penalty:.6f}  Feasible={ind.feasible}\n")
            f.write(f"Breakdown={ind.vio_breakdown}\n\n")
            for b in batches:
                key    = (b.origin, b.destination, b.batch_id)
                allocs = ind.od_allocations.get(key, [])
                if not allocs: continue
                f.write(f"Batch {b.batch_id}: {b.origin} -> {b.destination}, Q={b.quantity}\n")
                for a in allocs: f.write(str(a) + "\n")
                f.write("\n")
            f.write("\n")
    print(f"[EXPORT] {len(pareto)} Pareto solutions → {filename}")


def export_pareto_points_json(pareto, batches, out_json="pareto_points.json"):
    out = []
    for ind in pareto:
        sol = {
            "objectives": {
                "cost":          float(ind.objectives[0]),
                "emission_gCO2": float(ind.objectives[1]),
                "time_h":        float(ind.objectives[2]),
                "penalty":       float(ind.penalty),
            },
            "feasible":      bool(ind.feasible),
            "vio_breakdown": {k: float(v) for k, v in (ind.vio_breakdown or {}).items()},
            "border_flow":   {str(k): float(v) for k, v in getattr(ind, "border_flow", {}).items()},
            "border_util":   {str(k): float(v) for k, v in getattr(ind, "border_util", {}).items()},
            "allocations":   [],
        }
        for b in batches:
            key = (b.origin, b.destination, b.batch_id)
            blk = {
                "batch_id":    int(b.batch_id),
                "origin":      b.origin,
                "destination": b.destination,
                "quantity_teu": float(b.quantity),
                "ET":          float(b.ET),
                "LT":          float(b.LT),
                "paths":       [],
            }
            for a in ind.od_allocations.get(key, []):
                blk["paths"].append({
                    "share": float(a.share),
                    "nodes": list(a.path.nodes),
                    "modes": list(a.path.modes),
                    "base_cost_per_teu":     float(a.path.base_cost_per_teu),
                    "base_emission_per_teu": float(a.path.base_emission_per_teu),
                    "base_travel_time_h":    float(a.path.base_travel_time_h),
                })
            sol["allocations"].append(blk)
        out.append(sol)
    with open(out_json, "w", encoding="utf-8") as f:
        _json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"[EXPORT] pareto_points.json → {out_json}  ({len(out)} solutions)")


# ════════════════════════════════════════════════════════
# Plotting
# ════════════════════════════════════════════════════════

def _band(ax, gen, mean, std, color, alpha=0.18):
    ax.fill_between(gen, mean - std, mean + std, alpha=alpha, color=color)


def plot_pareto_3d(pts, save, title="NSGA-II Pareto Front 3D"):
    if not pts: return
    A = _finite_points_array(pts)
    if A.shape[0] == 0: return
    fig = plt.figure(figsize=(7, 6), dpi=200)
    ax  = fig.add_subplot(111, projection="3d")
    sc  = ax.scatter(A[:, 0], A[:, 1], A[:, 2], c=A[:, 0], cmap="viridis", s=40, alpha=0.9)
    plt.colorbar(sc, ax=ax, pad=0.1, fraction=0.04, label="Cost ($)")
    ax.set_xlabel("Cost ($)"); ax.set_ylabel("Emission (gCO₂)"); ax.set_zlabel("Time (h)")
    ax.set_title(title); ax.grid(True, ls=":", alpha=0.4)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_pareto_2d(pts, save):
    if not pts: return
    A = _finite_points_array(pts)
    if A.shape[0] == 0: return
    fig, axes = plt.subplots(1, 3, figsize=(14, 4.5), dpi=180)
    pairs = [(0,1,"Cost ($)","Emission (gCO₂)"),
             (0,2,"Cost ($)","Time (h)"),
             (1,2,"Emission (gCO₂)","Time (h)")]
    for ax, (xi, yi, xl, yl) in zip(axes, pairs):
        ax.scatter(A[:, xi], A[:, yi], s=30, alpha=0.85, c=A[:, 0], cmap="viridis")
        ax.set_xlabel(xl); ax.set_ylabel(yl)
        ax.set_title(f"{xl} vs {yl}")
        ax.grid(True, ls=":", alpha=0.45)
    fig.suptitle("NSGA-II Final Pareto Front — 2D Projections", fontsize=11, fontweight="bold")
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


# ════════════════════════════════════════════════════════
# Core NSGA-II runner
# ════════════════════════════════════════════════════════

def run_nsga2(
    node_names, node_region, node_hold_cost, node_proc_cost, node_trans_cost,
    arcs, timetables, batches,
    waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
    carbon_tax_map, emission_factor_map, mode_speeds_map,
    trans_map, border_delay_map, theta_rm, path_lib,
    pop_size=250, generations=200,
    archive_size=None,
):
    if archive_size is None:
        archive_size = pop_size

    tt_dict    = build_timetable_dict(timetables)
    arc_lookup = build_arc_lookup(arcs)

    eval_kwargs = dict(
        node_hold_cost=node_hold_cost,
        node_proc_cost=node_proc_cost,
        carbon_tax_map=carbon_tax_map,
        trans_map=trans_map,
        border_delay_map=border_delay_map,
        theta_rm=theta_rm,
        node_trans_cost=node_trans_cost,
    )

    # ── Initialise population: 1/6 each greedy direction, 1/2 random ─────
    n_each = max(1, pop_size // 6)
    population: List[Individual] = []
    for i in range(pop_size):
        if i < n_each:
            ind = greedy_initial_min_cost(batches, path_lib)
        elif i < 2 * n_each:
            ind = greedy_initial_individual(batches, path_lib)        # min-time
        elif i < 3 * n_each:
            ind = greedy_initial_min_emission(batches, path_lib)
        else:
            ind = random_initial_individual(batches, path_lib)
        repair_missing_allocations(ind, batches, path_lib)
        evaluate_individual(ind, batches, arcs, tt_dict,
                        waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
        population.append(ind)

    # Initial sort + crowding so first tournament has valid keys
    fronts0 = fast_non_dominated_sort(population)
    for front in fronts0:
        crowding_distance_assignment(population, front)

    front_hist_objs:            List[List[Tuple]] = []
    feasible_ratio_hist:        List[float]       = []
    feasible_ratio_strict_hist: List[float]       = []
    vio_mean_hist = {k: [] for k in ["miss_alloc","miss_tt","cap_excess",
                                     "late_teu_h","wait_teu_h"]}
    boost_trigger_hist:  List[int] = []
    boost_new_feas_hist: List[int] = []
    pareto_size_hist:    List[int] = []
    feasible_count_hist: List[int] = []

    _run_start = time.perf_counter()
    _prev_best = [float("inf")] * NUM_OBJ

    for gen in range(generations):

        # ── Step 1: Binary tournament → mating pool ──────
        mating_pool = [nsga2_binary_tournament(population) for _ in range(pop_size)]

        # ── Step 2: Crossover + Mutation → offspring Q ───
        offspring: List[Individual] = []
        while len(offspring) < pop_size:
            p1, p2 = random.sample(mating_pool, 2)
            if random.random() < CROSSOVER_RATE:
                c1, c2 = crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup)
            else:
                c1 = random_initial_individual(batches, path_lib)
                c2 = random_initial_individual(batches, path_lib)

            for child in (c1, c2):
                repair_missing_allocations(child, batches, path_lib)

            if random.random() < MUTATION_RATE:
                mutate_fixed(c1, batches, path_lib, tt_dict, arc_lookup,
                             arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                             **eval_kwargs)
            if random.random() < MUTATION_RATE:
                mutate_fixed(c2, batches, path_lib, tt_dict, arc_lookup,
                             arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                             **eval_kwargs)

            repair_missing_allocations(c1, batches, path_lib)
            repair_missing_allocations(c2, batches, path_lib)
            evaluate_individual(c1, batches, arcs, tt_dict,
                                waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
            evaluate_individual(c2, batches, arcs, tt_dict,
                                waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
            offspring.extend([c1, c2])
        offspring = offspring[:pop_size]

        # ── Step 3: Environmental selection (P ∪ Q → new P) ──
        combined   = population + offspring
        population = nsga2_environmental_selection(combined, pop_size)

        # ── Record metrics from population ───────────────
        feas_pop     = [ind for ind in population if ind.feasible]
        non_dom_feas = [ind for ind in population if ind.rank == 0 and ind.feasible]
        display      = unique_individuals_by_objectives(non_dom_feas, tol=1e-3)

        front_hist_objs.append([ind.objectives for ind in display])
        feas_total = len(feas_pop)
        feasible_ratio_hist.append(feas_total / max(1, pop_size))
        feasible_ratio_strict_hist.append(
            sum(1 for i in population if i.feasible_hard) / max(1, pop_size))
        pareto_size_hist.append(len(display))
        feasible_count_hist.append(feas_total)

        for k in vio_mean_hist:
            vals = [ind.vio_breakdown.get(k, 0.0) for ind in population]
            vio_mean_hist[k].append(float(np.mean(vals)))

        elapsed  = time.perf_counter() - _run_start
        if feas_pop:
            cur    = [min(i.objectives[j] for i in feas_pop) for j in range(NUM_OBJ)]
            d      = ["↓" if cur[j] < _prev_best[j] - 1e-3 else "→" for j in range(NUM_OBJ)]
            _prev_best = cur
            obj_str = (f"Cost={cur[0]:.3e}{d[0]} "
                       f"Emis={cur[1]:.3e}{d[1]} "
                       f"Time={cur[2]:.1f}h{d[2]}")
        else:
            obj_str = "No feasible solutions yet"

        best_pen = min(i.penalty for i in population) if population else float("inf")
        sep      = "=" * 72
        print(f"\n{sep}")
        print(f"  [NSGA-II] Gen {gen:03d}/{generations-1}  |  {elapsed:.1f}s elapsed")
        print(f"  Pop feasible: {feas_total}/{pop_size} "
              f"({feasible_ratio_hist[-1]:.1%})"
              f"  |  NonDom={len(display)}"
              f"  |  BestPenalty={best_pen:.2e}")
        print(f"  Best feasible: {obj_str}")
        print(sep)

        # ── Step 4: Feasibility boost ────────────────────
        boost_triggered = boost_new_feas = 0
        if feas_total < MIN_FEASIBLE_SOLUTIONS:
            boost_triggered = 1
            population, boost_new_feas = feasibility_boost(
                population, batches, path_lib, tt_dict, arc_lookup,
                arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h, eval_kwargs)
            # Re-rank after boost so next tournament uses fresh keys
            fronts_after = fast_non_dominated_sort(population)
            for front in fronts_after:
                crowding_distance_assignment(population, front)
            after = sum(1 for i in population if i.feasible)
            print(f"  ⚡ [BOOST] {feas_total} < {MIN_FEASIBLE_SOLUTIONS} "
                  f"→ after boost: {after} (+{boost_new_feas})")

        boost_trigger_hist.append(boost_triggered)
        boost_new_feas_hist.append(boost_new_feas)

    # ── Final Pareto: rank-0 feasible from population ────────
    fronts_final = fast_non_dominated_sort(population)
    for front in fronts_final:
        crowding_distance_assignment(population, front)

    pareto = unique_individuals_by_objectives(
        [ind for ind in population if ind.rank == 0 and ind.feasible], tol=1e-3)
    if not pareto:
        feas_remaining = [ind for ind in population if ind.feasible]
        if feas_remaining:
            pareto = unique_individuals_by_objectives(
                sorted(feas_remaining, key=lambda x: (x.rank, -x.crowding_distance)),
                tol=1e-3)

    total_t = time.perf_counter() - _run_start
    print(f"\n{'='*72}")
    print(f"  [NSGA-II] Run complete: {generations} gens, {total_t:.1f}s, Pareto={len(pareto)}")
    print(f"  ⚡ Boost: {sum(boost_trigger_hist)} gens triggered, "
          f"{sum(boost_new_feas_hist)} new feasible")
    print(f"{'='*72}")

    return (
        population, pareto,
        front_hist_objs,
        feasible_ratio_hist, feasible_ratio_strict_hist,
        vio_mean_hist,
        boost_trigger_hist, boost_new_feas_hist,
        pareto_size_hist, feasible_count_hist,
        total_t,
    )


# ════════════════════════════════════════════════════════
# Main  ── NSGA-II Tuned Baseline 30 runs
# ════════════════════════════════════════════════════════

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Run NSGA-II V1 baseline on an expanded multimodal network.")
    parser.add_argument("--data", default="data_expanded.xlsx", help="Input workbook path.")
    parser.add_argument("--pop", type=int, default=250, help="Population size.")
    parser.add_argument("--gens", type=int, default=200, help="Number of generations.")
    parser.add_argument("--runs", type=int, default=30, help="Number of random-seed runs.")
    parser.add_argument("--seed", type=int, default=1000, help="Base random seed.")
    parser.add_argument("--out", default="out_baseline", help="Output directory.")
    parser.add_argument("--expected-batches", type=int, default=0,
                        help="Optional exact batch count check; 0 disables the check.")
    args = parser.parse_args()

    DATA_FILE    = args.data
    POP_SIZE     = args.pop
    GENERATIONS  = args.gens
    RUNS         = args.runs
    BASE_SEED    = args.seed
    OUTPUT_DIR   = args.out

    FSPath(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    print("=" * 72)
    print("  NSGA-II  TUNED BASELINE  (3-objective)")
    print("  ▶ All hyper-parameters frozen from Stage 1 / 2 / 3 grid search")
    print("=" * 72)
    print(f"  GA      : pop={POP_SIZE}  gen={GENERATIONS}  runs={RUNS}")
    print(f"            pc={CROSSOVER_RATE}  pm={MUTATION_RATE}")
    print(f"  PathLib : DFS={DFS_MAX_PATHS_PER_OD}  TopK={PATHS_TOPK_PER_CRITERION}  "
          f"LibCap={PATH_LIB_CAP_TOTAL}")
    print(f"  HV      : MC samples={HV_SAMPLES}  ref={HV_REF_NORM}")
    print(f"  Data    : {DATA_FILE}")
    print(f"  Batches : expanded workbook batches  caps=Excel")
    print("=" * 72)

    print("\n[INIT] Loading network data...")
    (node_names, node_region,
     node_hold_cost, node_proc_cost, node_trans_cost,
     arcs, timetables, raw_batches,
     waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
     carbon_tax_map, emission_factor_map, mode_speeds_map,
     trans_map, border_delay_map, theta_rm) = load_network_from_extended(DATA_FILE)

    if args.expected_batches:
        assert len(raw_batches) == args.expected_batches, \
            f"Expected {args.expected_batches} batches, got {len(raw_batches)}"
    print(f"[INIT] Batches confirmed: {len(raw_batches)}")

    print("[INIT] Building path library (shared across all runs)...")
    tt_dict    = build_timetable_dict(timetables)
    arc_lookup = build_arc_lookup(arcs)
    random.seed(0); np.random.seed(0)
    path_lib = build_path_library(
        node_names, node_region, arcs, raw_batches, tt_dict, arc_lookup)
    sanity_check_path_lib(raw_batches, path_lib)

    all_front_hist: List = []
    all_run_rows:   List = []
    all_paretos:    List = []

    _hv_runs, _igd_runs, _sp_runs = [], [], []
    _fr_runs, _frs_runs            = [], []
    _mc_runs, _me_runs, _mt_runs  = [], [], []
    _psz_runs, _fc_runs            = [], []
    _boost_runs                    = []

    total_wall_start = time.perf_counter()

    for run_id in range(RUNS):
        seed = BASE_SEED + run_id
        random.seed(seed); np.random.seed(seed)

        print(f"\n{'='*72}")
        print(f"  RUN {run_id+1:02d}/{RUNS}  |  seed={seed}")
        print(f"{'='*72}")

        (population, pareto,
         front_hist_objs,
         feasible_ratio_hist, feasible_ratio_strict_hist,
         vio_mean_hist,
         boost_trigger_hist, boost_new_feas_hist,
         pareto_size_hist, feasible_count_hist,
         runtime_s) = run_nsga2(
            node_names, node_region, node_hold_cost, node_proc_cost, node_trans_cost,
            arcs, timetables, raw_batches,
            waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
            carbon_tax_map, emission_factor_map, mode_speeds_map,
            trans_map, border_delay_map, theta_rm, path_lib,
            pop_size=POP_SIZE, generations=GENERATIONS,
        )

        all_front_hist.append(front_hist_objs)
        all_paretos.append(pareto)
        _fr_runs.append(feasible_ratio_hist)
        _frs_runs.append(feasible_ratio_strict_hist)
        _psz_runs.append(pareto_size_hist)
        _fc_runs.append(feasible_count_hist)
        _boost_runs.append(boost_trigger_hist)

        def _bmin_run(idx):
            vals = []
            for gf in front_hist_objs:
                a = _finite_points_array(gf)
                vals.append(float(np.min(a[:, idx])) if a.shape[0] > 0 else np.nan)
            return _ffill_nan(np.array(vals))

        _mc_runs.append(_bmin_run(0))
        _me_runs.append(_bmin_run(1))
        _mt_runs.append(_bmin_run(2))

        all_run_rows.append({
            "run_id":                   run_id,
            "seed":                     seed,
            "runtime_s":                round(runtime_s, 2),
            "final_pareto_size":        len(pareto),
            "final_FeasRatio_soft":     float(feasible_ratio_hist[-1]),
            "final_FeasRatio_strict":   float(feasible_ratio_strict_hist[-1]),
            "boost_gens_triggered":     int(sum(boost_trigger_hist)),
            "boost_new_feasible_total": int(sum(boost_new_feas_hist)),
        })
        print(f"  [Run {run_id+1:02d}] Pareto={len(pareto)}  "
              f"FeasSoft={feasible_ratio_hist[-1]:.1%}  "
              f"Runtime={runtime_s:.1f}s")

    total_wall = time.perf_counter() - total_wall_start
    print(f"\n[INFO] All {RUNS} runs done in {total_wall:.1f}s")

    # ── Build P* ─────────────────────────────────────────
    P_star = build_P_star_fast(all_front_hist)
    if P_star:
        P_arr  = np.array(P_star, dtype=float)
        mins, maxs = np.min(P_arr, axis=0), np.max(P_arr, axis=0)
    else:
        mins, maxs = np.zeros(NUM_OBJ), np.ones(NUM_OBJ)

    hv_calc = HypervolumeCalculator(ref_point=HV_REF_NORM, num_samples=HV_SAMPLES,
                                     seed=HV_MC_SEED)
    Pn = normalize_points(P_star, mins, maxs) if P_star else []

    for r, front_hist_objs in enumerate(all_front_hist):
        hv_h, igd_h, sp_h = [], [], []
        last_hv = last_igd = last_sp = 0.0
        for gi, gf in enumerate(front_hist_objs):
            if gi % HV_EVERY == 0:
                pts     = [tuple(x) for x in _finite_points_array(gf)]
                An      = clip_points(normalize_points(pts, mins, maxs), HV_REF_NORM) if pts else []
                last_hv = hv_calc.calculate_points(An) if An else 0.0
            if gi % METRIC_EVERY == 0:
                pts      = [tuple(x) for x in _finite_points_array(gf)]
                An       = normalize_points(pts, mins, maxs) if pts else []
                last_igd = igd_plus(Pn, An) if (Pn and An) else float("inf")
                last_sp  = spacing_metric(An) if An else 0.0
            hv_h.append(last_hv)
            igd_h.append(last_igd)
            sp_h.append(last_sp)
        _hv_runs.append(hv_h)
        _igd_runs.append(igd_h)
        _sp_runs.append(sp_h)
        all_run_rows[r]["final_HV_norm"]  = float(hv_h[-1])
        all_run_rows[r]["final_IGD_plus"] = float(igd_h[-1]) if np.isfinite(igd_h[-1]) else None
        all_run_rows[r]["final_Spacing"]  = float(sp_h[-1])

    hv_mat  = np.array(_hv_runs,  dtype=float)
    igd_mat = np.array(_igd_runs, dtype=float)
    sp_mat  = np.array(_sp_runs,  dtype=float)
    fr_mat  = np.array(_fr_runs,  dtype=float)
    frs_mat = np.array(_frs_runs, dtype=float)
    mc_mat  = np.array(_mc_runs,  dtype=float)
    me_mat  = np.array(_me_runs,  dtype=float)
    mt_mat  = np.array(_mt_runs,  dtype=float)
    psz_mat = np.array(_psz_runs, dtype=float)
    fc_mat  = np.array(_fc_runs,  dtype=float)

    gen_arr   = np.arange(GENERATIONS)
    hv_mean,  hv_std  = np.mean(hv_mat,  0), np.std(hv_mat,  0)
    igd_mean, igd_std = np.mean(igd_mat, 0), np.std(igd_mat, 0)
    sp_mean,  sp_std  = np.mean(sp_mat,  0), np.std(sp_mat,  0)
    fr_mean,  fr_std  = np.mean(fr_mat,  0), np.std(fr_mat,  0)
    frs_mean, frs_std = np.mean(frs_mat, 0), np.std(frs_mat, 0)

    best_run_idx  = int(np.argmax(hv_mat[:, -1]))
    best_pareto   = all_paretos[best_run_idx]
    best_boost    = _boost_runs[best_run_idx]
    min_cost_best = mc_mat[best_run_idx]
    min_emis_best = me_mat[best_run_idx]
    min_time_best = mt_mat[best_run_idx]

    mc_mean, mc_std = np.nanmean(mc_mat, 0), np.nanstd(mc_mat, 0)
    me_mean, me_std = np.nanmean(me_mat, 0), np.nanstd(me_mat, 0)
    mt_mean, mt_std = np.nanmean(mt_mat, 0), np.nanstd(mt_mat, 0)

    p = lambda name: str(FSPath(OUTPUT_DIR) / name)

    # ── Per-generation metrics CSV ────────────────────────
    df_metrics = pd.DataFrame({"generation": gen_arr})
    df_metrics["HV_mean"] = hv_mean;  df_metrics["HV_std"] = hv_std
    df_metrics["IGD_plus_mean"] = igd_mean; df_metrics["IGD_plus_std"] = igd_std
    df_metrics["Spacing_mean"] = sp_mean;   df_metrics["Spacing_std"] = sp_std
    df_metrics["FeasRatio_soft_mean"]   = fr_mean;  df_metrics["FeasRatio_soft_std"]   = fr_std
    df_metrics["FeasRatio_strict_mean"] = frs_mean; df_metrics["FeasRatio_strict_std"] = frs_std
    df_metrics["MinCost_best"]  = min_cost_best
    df_metrics["MinEmis_best"]  = min_emis_best
    df_metrics["MinTime_best"]  = min_time_best
    df_metrics["MinCost_mean"]  = mc_mean; df_metrics["MinCost_std"]  = mc_std
    df_metrics["MinEmis_mean"]  = me_mean; df_metrics["MinEmis_std"]  = me_std
    df_metrics["MinTime_mean"]  = mt_mean; df_metrics["MinTime_std"]  = mt_std
    df_metrics["boost_triggered_best"] = best_boost
    df_metrics.to_csv(p("metrics_per_generation.csv"), index=False)
    print(f"[EXPORT] metrics_per_generation.csv saved.")

    # ── Per-run summary Excel ─────────────────────────────
    df_runs = pd.DataFrame(all_run_rows)
    def _s(col):
        v = df_runs[col].dropna().astype(float)
        return f"{v.mean():.4f} ± {v.std():.4f}"
    agg_row = {c: _s(c) if df_runs[c].dtype != object else "—" for c in df_runs.columns}
    agg_row["run_id"] = "mean±std"
    df_runs = pd.concat([df_runs, pd.DataFrame([agg_row])], ignore_index=True)
    xlsx_path = p("run_summary.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_runs.to_excel(writer, sheet_name="RunSummary", index=False)
        ws = writer.sheets["RunSummary"]
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            hfill = PatternFill("solid", fgColor="1F4E79")
            hfont = Font(color="FFFFFF", bold=True)
            afill = PatternFill("solid", fgColor="FFF2CC")
            for ci, col in enumerate(df_runs.columns, 1):
                c = ws.cell(row=1, column=ci)
                c.fill, c.font = hfill, hfont
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = min(len(col)+4, 24)
            for ci in range(1, len(df_runs.columns)+1):
                ws.cell(row=len(df_runs)+1, column=ci).fill = afill
            ws.freeze_panes = "A2"
        except Exception:
            pass
    print(f"[EXPORT] run_summary.xlsx → {xlsx_path}")

    # ── Pareto outputs (best run) ─────────────────────────
    save_pareto_solutions(best_pareto, raw_batches, p("result.txt"))
    export_pareto_points_json(best_pareto, raw_batches, out_json=p("pareto_points.json"))

    # ── Front history JSON ────────────────────────────────
    _fh_export = []
    for fh in all_front_hist:
        _fh_export.append([[list(obj) for obj in gf] for gf in fh])
    with open(p("front_hist_baseline.json"), "w", encoding="utf-8") as _f:
        _json.dump(_fh_export, _f)
    print(f"[EXPORT] nsga2_front_hist.json saved.")

    # ── Plots ─────────────────────────────────────────────
    ip_m = np.where(np.isfinite(igd_mean), igd_mean, np.nan)
    ip_s = np.where(np.isfinite(igd_std),  igd_std,  np.nan)

    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen_arr, hv_mean, lw=2.2, color="#9C27B0", label="HV mean")
    _band(ax, gen_arr, hv_mean, hv_std, "#9C27B0")
    ax.plot(gen_arr, hv_mat[best_run_idx], lw=1.4, ls="--",
            color="#E91E63", alpha=0.8, label=f"Best run #{best_run_idx}")
    ax.set_xlabel("Generation"); ax.set_ylabel("HV (normalised)")
    ax.set_ylim(bottom=0); ax.set_title(f"NSGA-II Hypervolume — mean±std ({RUNS} runs)")
    ax.legend(fontsize=8); ax.grid(True, ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_HV.png")); plt.close()

    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen_arr, ip_m, lw=2.2, color="#E53935", label="IGD+ mean")
    ax.fill_between(gen_arr,
                    np.where(np.isfinite(ip_m - ip_s), ip_m - ip_s, np.nan),
                    np.where(np.isfinite(ip_m + ip_s), ip_m + ip_s, np.nan),
                    alpha=0.18, color="#E53935")
    ax.set_xlabel("Generation"); ax.set_ylabel("IGD+ (lower is better)")
    ax.set_title(f"NSGA-II IGD+ — mean±std ({RUNS} runs)")
    ax.legend(fontsize=8); ax.grid(True, ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_IGDplus.png")); plt.close()

    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen_arr, sp_mean, lw=2.2, color="#00897B", label="Spacing mean")
    _band(ax, gen_arr, sp_mean, sp_std, "#00897B")
    ax.set_xlabel("Generation"); ax.set_ylabel("Spacing")
    ax.set_title(f"NSGA-II Spacing — mean±std ({RUNS} runs)")
    ax.legend(fontsize=8); ax.grid(True, ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_Spacing.png")); plt.close()

    fig, ax1 = plt.subplots(figsize=(10, 4), dpi=180)
    ax1.plot(gen_arr, hv_mean, lw=2.0, color="#9C27B0", label="HV mean")
    _band(ax1, gen_arr, hv_mean, hv_std, "#9C27B0")
    ax1.set_xlabel("Generation"); ax1.set_ylabel("HV (normalised)", color="#9C27B0")
    ax1.tick_params(axis="y", labelcolor="#9C27B0"); ax1.set_ylim(bottom=0)
    ax2 = ax1.twinx()
    ax2.plot(gen_arr, ip_m, lw=2.0, ls="--", color="#E53935", label="IGD+ mean")
    ax2.set_ylabel("IGD+", color="#E53935")
    ax2.tick_params(axis="y", labelcolor="#E53935")
    lines  = ax1.get_legend_handles_labels()[0] + ax2.get_legend_handles_labels()[0]
    labels = ax1.get_legend_handles_labels()[1] + ax2.get_legend_handles_labels()[1]
    ax1.legend(lines, labels, loc="lower right", fontsize=8)
    ax1.grid(True, ls=":", alpha=0.5)
    plt.title(f"NSGA-II Convergence: HV & IGD+ ({RUNS} runs)")
    plt.tight_layout(); plt.savefig(p("plot_Convergence.png")); plt.close()

    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen_arr, fr_mean,  lw=2.2, color="#1976D2", label="Soft mean")
    _band(ax, gen_arr, fr_mean, fr_std, "#1976D2")
    ax.plot(gen_arr, frs_mean, lw=2.0, ls="--", color="#FB8C00", label="Strict mean")
    _band(ax, gen_arr, frs_mean, frs_std, "#FB8C00", alpha=0.13)
    bgs = [g for g, v in enumerate(best_boost) if v > 0]
    if bgs:
        ax.scatter(bgs, [fr_mean[g] for g in bgs if g < len(fr_mean)],
                   marker="^", color="red", s=40, zorder=5,
                   label=f"Boost best run ({len(bgs)})")
    ax.set_ylim(-0.02, 1.05)
    ax.set_xlabel("Generation"); ax.set_ylabel("Feasible Ratio")
    ax.set_title(f"NSGA-II Feasible Ratio — mean±std ({RUNS} runs)")
    ax.legend(fontsize=8); ax.grid(True, ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_FeasibleRatio.png")); plt.close()

    fig, axes = plt.subplots(3, 1, figsize=(10, 9), dpi=180, sharex=True)
    fig.subplots_adjust(hspace=0.35)
    for ax, (mean, std, best, ylabel, color) in zip(axes, [
        (mc_mean, mc_std, min_cost_best, "Min Cost ($)",       "#E53935"),
        (me_mean, me_std, min_emis_best, "Min Emission (gCO₂)","#00897B"),
        (mt_mean, mt_std, min_time_best, "Min Time (h)",        "#FB8C00"),
    ]):
        ax.plot(gen_arr, mean, lw=2.0, color=color, label="Mean")
        _band(ax, gen_arr, mean, std, color)
        ax.plot(gen_arr, best, lw=1.4, ls="--", color="black",
                alpha=0.7, label=f"Best run #{best_run_idx}")
        mask = np.isfinite(best)
        if mask.any():
            fg = int(np.where(mask)[0][0])
            ax.axvline(fg, color="green", ls=":", lw=1.0, alpha=0.7)
        ax.set_ylabel(ylabel, fontsize=9); ax.legend(fontsize=7)
        ax.grid(True, ls=":", alpha=0.5)
    axes[-1].set_xlabel("Generation")
    fig.suptitle(f"NSGA-II Min Objectives per Generation ({RUNS} runs, mean±std)",
                 fontsize=12, fontweight="bold")
    plt.savefig(p("plot_MinObjectives.png")); plt.close()

    best_pts = unique_objective_tuples(
        [ind.objectives for ind in best_pareto if ind.feasible], tol=1e-9)
    plot_pareto_3d(best_pts, save=p("plot_Pareto3D.png"),
                   title=f"NSGA-II Pareto — Best Run #{best_run_idx}")
    plot_pareto_2d(best_pts, save=p("plot_Pareto2D.png"))

    all_pts_by_run = [[ind.objectives for ind in run_p if ind.feasible]
                      for run_p in all_paretos]
    fig, axes = plt.subplots(1, 3, figsize=(16, 5), dpi=180)
    pairs = [(0,1,"Cost ($)","Emission (gCO₂)"),
             (0,2,"Cost ($)","Time (h)"),
             (1,2,"Emission (gCO₂)","Time (h)")]
    for ax, (xi, yi, xl, yl) in zip(axes, pairs):
        for r_idx, pts in enumerate(all_pts_by_run):
            arr = _finite_points_array(pts)
            if arr.shape[0] == 0: continue
            if r_idx == best_run_idx: continue
            ax.scatter(arr[:, xi], arr[:, yi], s=10, alpha=0.3, color="#90CAF9")
        ab = _finite_points_array(all_pts_by_run[best_run_idx])
        if ab.shape[0] > 0:
            sc = ax.scatter(ab[:, xi], ab[:, yi], s=28, alpha=0.95,
                            c=ab[:, 0], cmap="viridis", zorder=5,
                            label=f"Best run #{best_run_idx}")
            plt.colorbar(sc, ax=ax, fraction=0.04, pad=0.02, label="Cost ($)")
            ax.legend(fontsize=7)
        ax.set_xlabel(xl); ax.set_ylabel(yl); ax.set_title(f"{xl} vs {yl}")
        ax.grid(True, ls=":", alpha=0.45)
    fig.suptitle(f"NSGA-II Pareto Front 2D — All {RUNS} Runs (grey) + Best Run (colour)",
                 fontsize=11, fontweight="bold")
    plt.tight_layout(); plt.savefig(p("plot_Pareto2D_allruns.png")); plt.close()

    psz_mean = np.mean(psz_mat, 0); psz_std = np.std(psz_mat, 0)
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen_arr, psz_mean, lw=2.2, color="#5C6BC0", label="Mean")
    _band(ax, gen_arr, psz_mean, psz_std, "#5C6BC0")
    ax.plot(gen_arr, psz_mat[best_run_idx], lw=1.4, ls="--",
            color="#E91E63", alpha=0.8, label=f"Best run #{best_run_idx}")
    ax.set_xlabel("Generation"); ax.set_ylabel("Non-dominated feasible size")
    ax.set_title(f"NSGA-II Pareto Size — mean±std ({RUNS} runs)")
    ax.legend(fontsize=8); ax.grid(True, ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_ParetoSize.png")); plt.close()

    fc_mean = np.mean(fc_mat, 0); fc_std = np.std(fc_mat, 0)
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen_arr, fc_mean, lw=2.2, color="#1976D2", label="Mean")
    _band(ax, gen_arr, fc_mean, fc_std, "#1976D2")
    ax.axhline(POP_SIZE, color="grey", ls="--", lw=1.0, label=f"Pop={POP_SIZE}")
    ax.set_xlabel("Generation"); ax.set_ylabel("Feasible solutions in population")
    ax.set_title(f"NSGA-II Feasible Count — mean±std ({RUNS} runs)")
    ax.legend(fontsize=8); ax.grid(True, ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_FeasibleCount.png")); plt.close()

    rts = [r["runtime_s"] for r in all_run_rows[:-1]]
    avg_rt = float(np.mean(rts))
    colors_rt = ["#E53935" if r > avg_rt else "#1976D2" for r in rts]
    fig, ax = plt.subplots(figsize=(max(8, RUNS*0.4), 4), dpi=180)
    ax.bar(range(RUNS), rts, color=colors_rt, alpha=0.85)
    ax.axhline(avg_rt, color="black", ls="--", lw=1.5, label=f"Mean={avg_rt:.1f}s")
    ax.set_xlabel("Run ID"); ax.set_ylabel("Runtime (s)")
    ax.set_title("NSGA-II Runtime per Run"); ax.legend()
    ax.grid(axis="y", ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(p("plot_Runtime.png")); plt.close()

    final_hv  = hv_mat[:,  -1]
    final_igd = igd_mat[:, -1]
    final_sp  = sp_mat[:,  -1]
    final_fr  = fr_mat[:,  -1]
    final_frs = frs_mat[:, -1]
    table_data = []
    for name, vals in [
        ("HV_norm (↑)",          final_hv),
        ("IGD+ (↓)",             final_igd),
        ("Spacing (↓)",          final_sp),
        ("FeasRatio soft (↑)",   final_fr),
        ("FeasRatio strict (↑)", final_frs),
        ("Runtime (s)",          np.array(rts)),
    ]:
        f = vals[np.isfinite(vals)]
        if len(f) == 0:
            table_data.append([name, "—", "—", "—", "—"])
        else:
            table_data.append([name,
                                f"{np.min(f):.4f}", f"{np.max(f):.4f}",
                                f"{np.mean(f):.4f}", f"{np.std(f):.4f}"])
    fig, ax = plt.subplots(figsize=(10, 3), dpi=150)
    ax.axis("off")
    tbl = ax.table(cellText=table_data,
                   colLabels=["Metric","Min","Max","Mean","Std"],
                   cellLoc="center", loc="center", bbox=[0,0,1,1])
    tbl.auto_set_font_size(False); tbl.set_fontsize(10)
    for j in range(5):
        tbl[(0,j)].set_facecolor("#1F4E79")
        tbl[(0,j)].set_text_props(color="white", fontweight="bold")
    for i in range(len(table_data)):
        clr = "#EBF3FB" if i % 2 == 0 else "#FFFFFF"
        for j in range(5): tbl[(i+1,j)].set_facecolor(clr)
    fig.suptitle(f"NSGA-II Tuned Baseline — {RUNS} Runs (Final Generation)",
                 fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout(); plt.savefig(p("plot_SummaryTable.png"),
                                    bbox_inches="tight"); plt.close()

    print("\n" + "=" * 72)
    print(f"  NSGA-II TUNED BASELINE  {RUNS} RUNS COMPLETE")
    print(f"  Total wall time:    {total_wall:.1f}s  ({total_wall/60:.1f} min)")
    print(f"  Best run:           #{best_run_idx}  (highest final HV)")
    print(f"  Pareto size (best): {len(best_pareto)}")
    print(f"  HV   mean±std:      {np.mean(final_hv):.4f} ± {np.std(final_hv):.4f}")
    fin_igd = final_igd[np.isfinite(final_igd)]
    if len(fin_igd):
        print(f"  IGD+ mean±std:      {np.mean(fin_igd):.4f} ± {np.std(fin_igd):.4f}")
    print(f"  Spacing mean±std:   {np.mean(final_sp):.4f} ± {np.std(final_sp):.4f}")
    print(f"  FeasSoft mean±std:  {np.mean(final_fr):.1%} ± {np.std(final_fr):.1%}")
    print(f"  Runtime mean±std:   {np.mean(rts):.1f}s ± {np.std(rts):.1f}s")
    print(f"  Outputs in:         {OUTPUT_DIR}/")
    print("=" * 72)
